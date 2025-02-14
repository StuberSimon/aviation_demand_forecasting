class DummyFile(object):
    def write(self, x): pass
    def flush(self): pass
import os
# Suppress TensorFlow warnings
os.environ['TF_CPP_MIN_LOG_LEVEL'] = '3'
import numpy as np
import pandas as pd
from joblib import dump, load
import tensorflow as tf
from tensorflow import keras
from keras import layers
from tqdm import tqdm
import random
from datetime import datetime, timedelta
from tsfresh.utilities.dataframe_functions import roll_time_series
from sklearn.preprocessing import MinMaxScaler
import openpyxl
import time
import contextlib
import shutil
import gc
import matplotlib.pylab as plt
import subprocess
import sys
dummy_file = DummyFile()

# Set random seeds for reproducibility
def set_random_seed(seed):
    np.random.seed(seed)
    tf.random.set_seed(seed)
    random.seed(seed)
    os.environ['PYTHONHASHSEED'] = str(seed)
    os.environ['TF_DETERMINISTIC_OPS'] = '1'

def movingAverage(array, size, average_season_features, first=True):
    # Convert array to DataFrame if it is not already one
    if not isinstance(array, pd.DataFrame):
        idx_temp = array.index
        array = pd.DataFrame(array)
        array.index = idx_temp
    if first:
        result = np.zeros(array.shape, dtype=float)
    else:
        result = np.zeros((1, array.shape[1]), dtype=float)
    for col in range(array.shape[1]):
        if first:
            if (not average_season_features) and (array.columns[col][:8] in ['3_months', '4_months', '6_months'] or array.columns[col][:9] in ['12_months', '17_months', '54_months']):
                    result[:, col] = array.iloc[:, col]
            else:
                for i in range(array.shape[0]):
                    if i < size + 1:
                        result[i, col] = np.mean(array.iloc[:i+1, col])
                    elif i >= array.shape[0] - size:
                        result[i, col] = np.mean(array.iloc[i-size:, col])
                    else:
                        result[i, col] = np.mean(array.iloc[i-size:i+1, col])
        else: # If not first
            if not average_season_features and (array.columns[col][:8] in ['3_months', '4_months', '6_months'] or array.columns[col][:9] in ['12_months', '17_months', '54_months']):
                result[0, col] = array.iloc[-1, col]
            else:
                result[0, col] = np.mean(array.iloc[-size-1:, col]) # Just last row
    if first:
        result_df = pd.DataFrame(result, columns=array.columns, index=array.index)
    else:
        result_df = pd.DataFrame(result, columns=array.columns)
    return result_df

def main():    
    # Base path for the extracted features with '/'
    base_path = '01_extracted_features/'
    models_path = '02_models/'
    summary_path = 'Summary.xlsx'
    #sbs_path = '02_models/01_models_step_by_step/'
    sbs_path = models_path
    plot_path = '03_plots/'

    model_ids = [2182,2183,2184,2185,2186,2187,2188,2189,2190,2191,2192,2193,2194,2195,2196,2197,2198,2199,2200,2201,2202,2203,2204,2205,2206,2207,2208,2233,2234,2235,2236,2237,2238,2239,2240,2241,2242,2243,2244,2245,2246,2247,2248,2249,2250,2251,2252,2253,2254,2255,2256,2257,2258,2259]

    learning_rates_list = [1e-7, 5e-7, 1e-8]
    num_epochs_list = [1, 2, 4, 8]

    seed_start = 42
    season_features = load(base_path + 'season_features_3_4_6_12_17_54.joblib')
    feature_importance = load(base_path + '249_feature_importance_df.joblib')
    gdp_features = load(base_path + 'gdp_features_v1.joblib').reset_index(drop=True)
    train_models_sbs(model_ids, learning_rates_list, num_epochs_list, seed_start, season_features, feature_importance, gdp_features, base_path, models_path, summary_path, sbs_path, plot_path)

def train_models_sbs(model_ids, learning_rates_list=None, num_epochs_list=None, seed_start=42, season_features=None, feature_importance=None, gdp_features=None, base_path=None, models_path=None, summary_path=None, sbs_path=None, plot_path=None):

    """
    Trains the models with covid data with specified learning rates to analyze the impact of including covid data.

    Args:
        model_ids (list<int>): List of model IDs to train.
        learning_rates_list (list<float>): List of learning rates to train the models with.
        num_epochs_list (list<int>): List of number of epochs to train the models with.
        seed_start (int): The starting seed for the models.
        season_features (pd.DataFrame): DataFrame with season features.
        feature_importance (pd.DataFrame): DataFrame with feature importance.
        gdp_features (pd.DataFrame): DataFrame with GDP features.
        base_path (str): Path to the extracted features.
        models_path (str): Path to the models folder.
        summary_path (str): Path to the summary file.
        sbs_path (str): Path to the folder to save the sbs models in.
        plot_path (str): Path to the plots folder.

    Returns:
        list<int>: List of model IDs that were trained.
    """



    # Set missing values to default
    learning_rates_list = learning_rates_list or [1e-7]
    num_epochs_list = num_epochs_list or [1]
    base_path = base_path or '01_extracted_features/'
    models_path = models_path or '02_models/'
    summary_path = summary_path or 'Summary.xlsx'
    plot_path = plot_path or '03_plots/'
    sbs_path = sbs_path or models_path
    season_features = season_features or load(base_path + 'season_features_3_4_6_12_17_54.joblib')
    feature_importance = feature_importance or load(base_path + '1290_feature_importance_df.joblib')
    gdp_features = gdp_features or load(base_path + 'gdp_features_v1.joblib').reset_index(drop=True)
    num_seeds = 1

    # Backup Summary
    # Get the base name of the source file
    base_name = os.path.basename(summary_path)
    # Get the file name and extension
    file_name, file_ext = os.path.splitext(base_name)
    # Get the current date
    current_date = datetime.now().strftime('%d.%m.%Y_%H-%M-%S')    
    # Create the new file name with date
    new_file_name = f"{file_name}_{current_date}{file_ext}"
    # Create the destination path
    dest_path = os.path.join('04_Summary_backups', new_file_name)
    # Copy the file to the new destination
    shutil.copy(summary_path, dest_path)

    # Build df for all models
    models_all = pd.DataFrame(
        columns=['ref_id', 'name', 'lookahead', 'data_start', 'data_end', 'learning_rate', 'train_loss',
                'val_loss', 'file_name', 'model', 'data_idx', 'best_seed',
                'time_frame_size', 'patience', 'min_delta', 'input_size',
                'season_features_size', 'feature_threshold', 'learning_rate_scheduler', 'include_flights',
                'num_roll_features', 'time_frame_features_size', 'time_frame_flights_size', 'average_window_size',
                'average_season_features', 'num_epochs', 'skip_me', 'ref_HMAPE', 'gdp_size'],
        index=range(len(model_ids) * len(learning_rates_list) * len(num_epochs_list))
    )

    for i in range(len(models_all)):
        models_all.loc[i, 'ref_id'] = model_ids[i % len(model_ids)]
        models_all.loc[i, 'learning_rate'] = learning_rates_list[(i // len(model_ids)) % len(learning_rates_list)]
        models_all.loc[i, 'num_epochs'] = num_epochs_list[(i // len(model_ids) // len(learning_rates_list)) % len(num_epochs_list)]


    workbook = openpyxl.load_workbook(summary_path)
    sheet = workbook.active

    for i in range(len(models_all)):
        row = None
        for r in sheet.iter_rows(min_row=1, max_row=sheet.max_row, values_only=True):
            if r[0] == models_all.loc[i, 'ref_id']:
                row = r
                break
        if row is None:
            print(f'Model {models_all.loc[i, "ref_id"]} not found in Summary.')
            continue
        models_all.loc[i, 'ref_id'] = row[0]
        models_all.loc[i, 'lookahead'] = row[2]
        models_all.loc[i, 'data_start'] = row[3]
        models_all.loc[i, 'data_end'] = row[4]
        models_all.loc[i, 'file_name'] = row[8]
        models_all.loc[i, 'name'] = 'SBS ' + row[1]
        models_all.loc[i, 'time_frame_size'] = row[9]
        models_all.loc[i, 'patience'] = row[10]
        models_all.loc[i, 'min_delta'] = row[11]
        models_all.loc[i, 'input_size'] = row[12]
        models_all.loc[i, 'season_features_size'] = row[18]
        models_all.loc[i, 'feature_threshold'] = row[19]
        models_all.loc[i, 'learning_rate_scheduler'] = row[20]
        models_all.loc[i, 'include_flights'] = row[21]
        models_all.loc[i, 'num_roll_features'] = row[22]
        models_all.loc[i, 'time_frame_features_size'] = row[23]
        models_all.loc[i, 'time_frame_flights_size'] = row[24]
        models_all.loc[i, 'average_window_size'] = row[25]
        models_all.at[i, 'best_seed'] = np.array([row[13]] * (row[4] - row[3]))# Best seeds from base model
        models_all.at[i, 'train_loss'] = np.array([row[6]] * (row[4] - row[3])) # Train loss from base model
        models_all.loc[i, 'average_season_features'] = row[27]
        models_all.loc[i, 'ref_HMAPE'] = row[30]
        models_all.loc[i, 'gdp_size'] = row[33]
        models_all.loc[i, 'add_timestamp'] = row[34]

    models_all = models_all[models_all['name'].notna()].reset_index(drop=True)

    workbook = openpyxl.load_workbook(summary_path)
    sheet = workbook.active
    rows_buffer = list()
    
    for i in range(len(models_all)):
        for row in sheet.iter_rows(min_row=1, values_only=True):
            model_in_summary = row[1] == models_all.loc[i, 'name'] and row[2] == models_all.loc[i, 'lookahead'] and row[3] == models_all.loc[i, 'data_start']
            model_in_summary = model_in_summary and row[4] == models_all.loc[i, 'data_end'] and row[5] == models_all.loc[i, 'learning_rate'] 
            model_in_summary = model_in_summary and row[8] == models_all.loc[i, 'file_name'] and row[9] == models_all.loc[i, 'time_frame_size']
            model_in_summary = model_in_summary and row[10] == models_all.loc[i, 'patience'] and row[11] == models_all.loc[i, 'min_delta']
            model_in_summary = model_in_summary and row[12] == models_all.loc[i, 'input_size']
            model_in_summary = model_in_summary and row[18] == models_all.loc[i, 'season_features_size']
            model_in_summary = model_in_summary and row[19] == models_all.loc[i, 'feature_threshold']
            model_in_summary = model_in_summary and row[20] == models_all.loc[i, 'learning_rate_scheduler']
            model_in_summary = model_in_summary and row[21] == models_all.loc[i, 'include_flights']
            model_in_summary = model_in_summary and row[22] == models_all.loc[i, 'num_roll_features']
            model_in_summary = model_in_summary and row[23] == models_all.loc[i, 'time_frame_features_size']
            model_in_summary = model_in_summary and row[24] == models_all.loc[i, 'time_frame_flights_size']
            model_in_summary = model_in_summary and row[25] == models_all.loc[i, 'average_window_size']
            model_in_summary = model_in_summary and row[28] == models_all.loc[i, 'ref_id']
            model_in_summary = model_in_summary and row[29] == models_all.loc[i, 'num_epochs']
            model_in_summary = model_in_summary and row[27] == models_all.loc[i, 'average_season_features']
            model_in_summary = model_in_summary and row[33] == models_all.loc[i, 'gdp_size']
            model_in_summary = model_in_summary and row[34] == models_all.loc[i, 'add_timestamp']

            if model_in_summary:
                print(f'Model {i} already in Summary with ID {row[0]}.')
                break
        models_all.loc[i, 'skip_me'] = model_in_summary    

    print(f'Models finished: {sum(models_all["skip_me"])}/{len(models_all)}.')
    models_all = models_all[models_all['skip_me'] == False].sort_values(by='ref_HMAPE', ascending=True).reset_index(drop=True)
    # Save models_all in csv
    models_all.to_excel('sbs_models_all.xlsx', index=True)
    print('Models_all saved.')

    start_time = time.time()
    total_iterations = len(models_all)
    ids = list()
    
    for i in range(len(models_all)):
        finished = False
        n_steps = 0
        data_end_original = models_all.loc[i, 'data_end']
        models_all.loc[i, 'data_end'] += 1
        predicted_values = np.array([])
        train_features_saved_avg = pd.DataFrame()
        train_labels_train_saved_avg = pd.DataFrame()

        while n_steps < 1000:
            if n_steps == 1:
                pbar_outer = tqdm(total=321-data_end_original, desc=f'SBS for Model {i}/{len(models_all)-1}')
            elif n_steps == 0:
                pbar_outer = None

            features_raw = pd.read_csv(base_path + models_all.loc[i, 'file_name']).drop(['YearMonth'], axis=1)

            actual_values = features_raw['Flights'].copy()

            important_features = feature_importance[feature_importance['cumulative_importance'] <= models_all.loc[i, 'feature_threshold']]
            important_features = important_features.sort_values(by='cumulative_importance', ascending=True)
            valid_features = [feature for feature in important_features['feature'] if feature in features_raw.columns]
            valid_features.append('Flights')
            features_raw = features_raw[valid_features]
            

            if models_all.loc[i, 'season_features_size'] > 0:
                features_raw = pd.concat([features_raw, season_features], axis=1) 
            if models_all.loc[i, 'gdp_size'] > 0:
                features_raw = pd.concat([gdp_features.iloc[:,:models_all.loc[i, 'gdp_size']], features_raw], axis=1)
            if models_all.loc[i, 'add_timestamp'] == 1:
                features_raw['Timestamp'] = features_raw.index

            lookahead = models_all.loc[i, 'lookahead']
            features_raw = features_raw.iloc[models_all.loc[i, 'data_start']:models_all.loc[i, 'data_end']]

            train_labels = features_raw['Flights'].shift(-lookahead).dropna()
            features_original = features_raw.copy()
            features_original['labels'] = features_raw['Flights']
            if lookahead > 0:
                if not finished:
                    features_raw = features_raw.iloc[:-lookahead]
                features_raw['labels'] = train_labels
                features_raw.reset_index(drop=True, inplace=True)
            train_dataset = features_raw.copy()

            idx = train_dataset.index

            # Normalize the data
            if n_steps == 0:
                scalers_dict = load(models_path + f'{models_all.loc[i, 'ref_id']}_scalers.joblib')
            for column in train_dataset.columns:
                scaler = scalers_dict[column]
                train_dataset[column] = scaler.transform(train_dataset[column].values.reshape(-1, 1))
                features_original[column] = scaler.transform(features_original[column].values.reshape(-1, 1))

            if models_all.loc[i, 'include_flights']:
                features_original = features_original.drop(['labels'], axis=1)
            else:
                features_original = features_original.drop(['labels', 'Flights'], axis=1)

            lookahead = models_all.loc[i, 'lookahead']
            time_frame_features_size = models_all.loc[i, 'time_frame_features_size'] - 1
            time_frame_flights_size = models_all.loc[i, 'time_frame_flights_size'] - 1
            time_frame_size = models_all.loc[i, 'time_frame_size']
            num_roll_features = models_all.loc[i, 'num_roll_features']

            if time_frame_size > 0 and len(train_dataset) >= time_frame_size:
                # Select first ten columns from train_dataset and append 'Flights' if not already in
                if time_frame_features_size > 0:
                    selected_columns = train_dataset.copy().drop(['labels', 'Flights'], axis=1).columns[:num_roll_features].tolist()
                else:
                    selected_columns = []
                # Train features
                if n_steps == 0: # First rolling
                    train_features_saved = pd.DataFrame()
                    if models_all.loc[i, 'include_flights'] and time_frame_flights_size == 0:
                        train_features = train_dataset.copy().drop(columns=selected_columns).drop(['labels'], axis=1)[time_frame_size:].reset_index(drop=True)
                    elif not models_all.loc[i, 'include_flights']:
                        train_features = train_dataset.copy().drop(columns=selected_columns).drop(['labels', 'Flights'], axis=1)[time_frame_size:].reset_index(drop=True)
                    else:
                        selected_columns.append('Flights')
                        train_features = train_dataset.copy().drop(columns=selected_columns, axis=1).drop(['labels'], axis=1)[time_frame_size:].reset_index(drop=True)
                else: # Select only the last row
                    if models_all.loc[i, 'include_flights'] and time_frame_flights_size == 0:
                        train_features = train_dataset.copy().drop(columns=selected_columns).drop(['labels'], axis=1)[time_frame_size:].reset_index(drop=True).tail(1)
                    elif not models_all.loc[i, 'include_flights']:
                        train_features = train_dataset.copy().drop(columns=selected_columns).drop(['labels', 'Flights'], axis=1)[time_frame_size:].reset_index(drop=True).tail(1)
                    else:
                        selected_columns.append('Flights')
                        train_features = train_dataset.copy().drop(columns=selected_columns, axis=1).drop(['labels'], axis=1)[time_frame_size:].reset_index(drop=True).tail(1)
                
                if n_steps == 0:
                    pbar_inner = tqdm(total=len(selected_columns), desc=f'Rolling train for model {i}/{len(models_all)-1}')
                else:
                    pbar_inner = None
                for feature in selected_columns:
                    if n_steps == 0:
                        train_dataset_melted = train_dataset[[feature]].copy()
                        train_dataset_melted['Timestamp'] = train_dataset.index
                        train_dataset_melted['Symbols'] = 'A'
                        # Roll time series
                        with contextlib.redirect_stdout(dummy_file), contextlib.redirect_stderr(dummy_file):
                            if feature == 'Flights' and time_frame_flights_size > 0 and models_all.loc[i, 'include_flights']:
                                train_features_rolled = roll_time_series(train_dataset_melted, column_id='Symbols',
                                                                        column_sort='Timestamp', max_timeshift=time_frame_flights_size,
                                                                        min_timeshift=time_frame_flights_size)
                            elif feature != 'Flights' and time_frame_features_size > 0:
                                train_features_rolled = roll_time_series(train_dataset_melted, column_id='Symbols',
                                                                        column_sort='Timestamp', max_timeshift=time_frame_features_size,
                                                                        min_timeshift=time_frame_features_size)
                            else:
                                continue
                            # Transform rolled features
                            grouped = train_features_rolled.groupby('id')
                            dfs = {name: group for name, group in grouped}
                            train_features_temp = pd.DataFrame()
                            for key in dfs.keys():
                                df = dfs[key].drop(['id', 'Symbols'], axis=1)
                                df.index = df['Timestamp'].values
                                df = df.drop(['Timestamp'], axis=1)
                                idx_temp = df.index[0]
                                # Transpose the DataFrame
                                df_reshaped = df.T
                                # Reset the index to ensure the original index is preserved
                                df_reshaped.columns = [f'{feature}_{i}' for i in range(0, len(df_reshaped.columns))]
                                df_reshaped.index = [idx_temp]
                                train_features_temp = pd.concat([train_features_temp, df_reshaped], axis=0)
                            if n_steps == 0 and feature == 'Flights':
                                if time_frame_flights_size < time_frame_features_size:
                                    train_features_temp = train_features_temp[time_frame_features_size-time_frame_flights_size:].reset_index(drop=True)
                            elif n_steps == 0 and feature != 'Flights':
                                if time_frame_features_size < time_frame_flights_size:
                                    train_features_temp = train_features_temp[time_frame_flights_size-time_frame_features_size:].reset_index(drop=True)
                        if pbar_inner:
                            pbar_inner.update(1)
                    else: # Rolling should only output one row
                        if feature == 'Flights' and time_frame_flights_size > 0 and models_all.loc[i, 'include_flights']:
                            train_dataset_melted = train_dataset[[feature]][-time_frame_flights_size-1:].copy().reset_index(drop=True)
                        elif feature != 'Flights' and time_frame_features_size > 0:
                            train_dataset_melted = train_dataset[[feature]][-time_frame_features_size-1:].copy().reset_index(drop=True)
                        else:
                            continue
                        # Ensure the input is a NumPy array
                        arr = np.asarray(train_dataset_melted)                            
                        # Create column names
                        col_names = [f"{feature}_{i}" for i in range(arr.shape[0])]                            
                        # Convert the array to a DataFrame and transpose it
                        train_features_temp = pd.DataFrame(arr).T                            
                        # Set the column names
                        train_features_temp.columns = col_names
                        train_features_temp.index = train_features.index
                        
                    train_features = pd.concat([train_features, train_features_temp], axis=1)
                train_features = pd.concat([train_features_saved, train_features], axis=0)
                train_labels = train_dataset.loc[train_features.index]['labels']
                train_features_saved = train_features.copy()
            else:
                if models_all.loc[i, 'include_flights']:
                    train_features = train_dataset.copy().drop(['labels'], axis=1)
                else:
                    train_features = train_dataset.copy().drop(['labels', 'Flights'], axis=1)

                train_labels = train_dataset['labels'].dropna()

            input_size = train_features.shape[1]
            models_all.loc[i, 'input_size'] = input_size
            # Check if model already in Summary
            workbook = openpyxl.load_workbook(summary_path)
            sheet = workbook.active
            model_in_summary = False
            less_seeds_row = 0
            less_seeds_best = seed_start
            less_seeds = False
            n = 0
            for row in sheet.iter_rows(min_row=1, values_only=True):
                model_in_summary = row[1] == models_all.loc[i, 'name'] and row[2] == models_all.loc[i, 'lookahead'] and row[3] == models_all.loc[i, 'data_start']
                model_in_summary = model_in_summary and row[4] == models_all.loc[i, 'data_end'] and row[5] == models_all.loc[i, 'learning_rate'] 
                model_in_summary = model_in_summary and row[8] == models_all.loc[i, 'file_name'] and row[9] == models_all.loc[i, 'time_frame_size']
                model_in_summary = model_in_summary and row[10] == models_all.loc[i, 'patience'] and row[11] == models_all.loc[i, 'min_delta']
                model_in_summary = model_in_summary and row[12] == input_size
                model_in_summary = model_in_summary and row[18] == models_all.loc[i, 'season_features_size']
                model_in_summary = model_in_summary and row[19] == models_all.loc[i, 'feature_threshold']
                model_in_summary = model_in_summary and row[20] == models_all.loc[i, 'learning_rate_scheduler']
                model_in_summary = model_in_summary and row[21] == models_all.loc[i, 'include_flights']
                model_in_summary = model_in_summary and row[22] == models_all.loc[i, 'num_roll_features']
                model_in_summary = model_in_summary and row[23] == models_all.loc[i, 'time_frame_features_size']
                model_in_summary = model_in_summary and row[24] == models_all.loc[i, 'time_frame_flights_size']
                model_in_summary = model_in_summary and row[25] == models_all.loc[i, 'average_window_size']
                model_in_summary = model_in_summary and row[28] == models_all.loc[i, 'ref_id']
                model_in_summary = model_in_summary and row[29] == models_all.loc[i, 'num_epochs']
                model_in_summary = model_in_summary and row[27] == models_all.loc[i, 'average_season_features']
                model_in_summary = model_in_summary and row[33] == models_all.loc[i, 'gdp_size']
                model_in_summary = model_in_summary and row[34] == models_all.loc[i, 'add_timestamp']
                n += 1
            if model_in_summary:
                print('something went wrong')
                continue
            workbook.close()
            best_history = None
            best_seed = None
            if models_all['name'].loc[i][:4] == 'LSTM':
                train_features = train_features.values.reshape(train_features.shape[0], 1, train_features.shape[1])
            if models_all.loc[i, 'average_window_size'] > 0:
                train_features_temp = movingAverage(train_features, models_all.loc[i, 'average_window_size'], models_all.loc[i, 'average_season_features'], n_steps==0)
                train_labels_train_temp = movingAverage(train_labels, models_all.loc[i, 'average_window_size'], models_all.loc[i, 'average_season_features'], n_steps==0)
                train_features = pd.concat([train_features_saved_avg, train_features_temp], axis=0).reset_index(drop=True)
                train_labels_train = pd.concat([train_labels_train_saved_avg, train_labels_train_temp], axis=0).reset_index(drop=True)
                train_features_saved_avg = train_features.copy()
                train_labels_train_saved_avg = train_labels_train.copy()
            else:
                train_labels_train = train_labels
            if n_steps == 0:
                train_features_original = train_features[:-1].copy()            
                print(f'Shape train labels: {train_labels.shape}\nShape train features: {train_features.shape}')
                print(f'Time: {datetime.now().strftime('%H:%M')}')
                model = load(models_path + f'{models_all.loc[i, 'ref_id']}_model.joblib')
            else:
                model = models_all.loc[i, 'model']
            if not finished and models_all.loc[i, 'data_end'] <= 309: # This is the index of the last data point during covid
                for j in range(less_seeds_best-seed_start,num_seeds):
                    with contextlib.redirect_stdout(dummy_file), contextlib.redirect_stderr(dummy_file):
                        set_random_seed(seed_start + j)
                        models_all.loc[i, 'model'] = model
                        # Update the learning rate
                        optimizer = models_all.loc[i, 'model'].optimizer
                        # Debug print statement to check the type and value of learning_rate
                        learning_rate = tf.constant(float(models_all.loc[i, 'learning_rate']), dtype=tf.float32)
                        optimizer.learning_rate.assign(learning_rate)

                        history = models_all.loc[i, 'model'].fit(
                            train_features.iloc[-1,:].values.reshape(1, -1), np.array([train_labels_train.iloc[-1]]),
                            epochs=models_all.loc[i, 'num_epochs'],
                            batch_size=1,
                            verbose=0
                        )
                        if best_history is None or history.history['loss'][0] < best_history.history['loss'][0]:
                            best_history = history
                            best_seed = seed_start + j
                            best_model = models_all.loc[i, 'model']
                models_all.at[i, 'train_loss'] = np.append(models_all.loc[i, 'train_loss'], best_history.history['loss'][0])
                models_all.at[i, 'best_seed'] = np.append(models_all.loc[i, 'best_seed'], best_seed)
                models_all.at[i, 'model'] = best_model
                del best_model
                del model
                del best_history
                del history
                gc.collect()
            elif finished: # If finished
                # Predict
                with contextlib.redirect_stdout(dummy_file), contextlib.redirect_stderr(dummy_file):
                    predicted_value = np.round(scalers_dict['labels'].inverse_transform(models_all.loc[i, 'model'].predict(train_features.iloc[-1,:].values.reshape(1, -1))).flatten())
                predicted_values = np.append(predicted_values, predicted_value)
                if pbar_outer:
                    pbar_outer.update(1)
                break

            if models_all.loc[i, 'data_end'] < 321: # This is the index of the last data point of the dataset
                n_steps += 1
            else:
                finished = True
            models_all.loc[i, 'data_end'] += 1
            # Predict
            with contextlib.redirect_stdout(dummy_file), contextlib.redirect_stderr(dummy_file):
                predicted_value = np.round(scalers_dict['labels'].inverse_transform(models_all.loc[i, 'model'].predict(train_features.iloc[-1,:].values.reshape(1, -1))).flatten())
            predicted_values = np.append(predicted_values, predicted_value)
            if pbar_outer:
                pbar_outer.update(1)
        
        # Compute MAPE
        # Split train and test predictions
        predicted_train = pd.DataFrame(predicted_values)[-20:-13].values.reshape(7,)
        predicted_test = pd.DataFrame(predicted_values)[-13:-1].values.reshape(12,)

        mape = np.mean(np.abs((actual_values[-19:-12].values - predicted_train) / actual_values[-19:-12].values)) * 100
        mape_test = np.mean(np.abs((actual_values[-12:].values - predicted_test) / actual_values[-12:].values)) * 100
        # Compute HMAPE
        hmape = 1/(7) * np.sum(np.abs(actual_values[-19:-12].values - predicted_train) /
                                    np.sqrt(np.abs(actual_values[-19:-12].values - predicted_train) + 1))
        hmape_test = 1/(12) * np.sum(np.abs(actual_values[-12:].values - predicted_test) /
                                    np.sqrt(np.abs(actual_values[-12:].values - predicted_test) + 1))

        

        model_original = load(models_path + f'{models_all.loc[i, 'ref_id']}_model.joblib')
        scalers_dict_original = load(models_path + f'{models_all.loc[i, 'ref_id']}_scalers.joblib')
        #features_original_ref = load(models_path + f'{models_all.loc[i, 'ref_id']}_features_original.joblib')
        with contextlib.redirect_stdout(dummy_file), contextlib.redirect_stderr(dummy_file):
            predicted_values_original = np.round(scalers_dict_original['labels'].inverse_transform(model_original.predict(train_features_original).reshape(-1, 1)).flatten()) #train_features[:data_end_original-lookahead-time_frame_size-1]
        predicted_values = np.append(predicted_values_original, predicted_values)




        # Save model in summary
        workbook = openpyxl.load_workbook(summary_path)
        sheet = workbook.active   

        # Add buffered rows to the sheet
        for row_data in rows_buffer:
            # Save model in summary
            if row_data['less_seeds']:
                last_row = row_data['less_seeds_row']
            else:
                # Find the last row with data in column 1
                last_row = 0
                for row in range(sheet.max_row, 0, -1):
                    if sheet.cell(row=row, column=1).value is not None:
                        last_row = row
                        break
            sheet.cell(row=last_row + 1, column=1, value=row_data['id'])
            sheet.cell(row=last_row + 1, column=2, value=row_data['name'])
            sheet.cell(row=last_row + 1, column=3, value=row_data['lookahead'])
            sheet.cell(row=last_row + 1, column=4, value=row_data['data_start'])
            sheet.cell(row=last_row + 1, column=5, value=row_data['data_end'])
            sheet.cell(row=last_row + 1, column=6, value=row_data['learning_rate'])
            sheet.cell(row=last_row + 1, column=9, value=row_data['file_name'])
            sheet.cell(row=last_row + 1, column=10, value=row_data['time_frame_size'])
            sheet.cell(row=last_row + 1, column=11, value=row_data['patience'])
            sheet.cell(row=last_row + 1, column=12, value=row_data['min_delta'])
            sheet.cell(row=last_row + 1, column=13, value=row_data['input_size'])
            sheet.cell(row=last_row + 1, column=15, value=row_data['num_seeds'])
            sheet.cell(row=last_row + 1, column=16, value=row_data['time'])
            sheet.cell(row=last_row + 1, column=16).number_format = 'DD.MM.YYYY HH:MM:SS'
            sheet.cell(row=last_row + 1, column=19, value=row_data['season_features_size'])
            sheet.cell(row=last_row + 1, column=20, value=row_data['feature_threshold'])
            sheet.cell(row=last_row + 1, column=21, value=row_data['learning_rate_scheduler'])
            sheet.cell(row=last_row + 1, column=22, value=row_data['include_flights'])
            sheet.cell(row=last_row + 1, column=23, value=row_data['num_roll_features'])
            sheet.cell(row=last_row + 1, column=24, value=row_data['time_frame_features_size'])
            sheet.cell(row=last_row + 1, column=25, value=row_data['time_frame_flights_size'])
            sheet.cell(row=last_row + 1, column=26, value=row_data['average_window_size'])
            sheet.cell(row=last_row + 1, column=29, value=row_data['ref_id'])
            sheet.cell(row=last_row + 1, column=27, value=row_data['mape'])
            sheet.cell(row=last_row + 1, column=30, value=row_data['num_epochs'])
            sheet.cell(row=last_row + 1, column=28, value=row_data['average_season_features'])
            sheet.cell(row=last_row + 1, column=31, value=row_data['hmape'])
            sheet.cell(row=last_row + 1, column=32, value=row_data['mape_test'])
            sheet.cell(row=last_row + 1, column=33, value=row_data['hmape_test'])
            sheet.cell(row=last_row + 1, column=34, value=row_data['gdp_size'])
            sheet.cell(row=last_row + 1, column=35, value=row_data['add_timestamp'])
        if less_seeds:
            last_row = less_seeds_row
        else:
            # Find the last row with data in column 1
            last_row = 0
            for row in range(sheet.max_row, 0, -1):
                if sheet.cell(row=row, column=1).value is not None:
                    last_row = row
                    break
        max_id = 0
        for row in sheet.iter_rows(min_row=1, values_only=True):
            if row[0] is not None and isinstance(row[0], int):
                max_id = max(max_id, row[0])
        id = max_id + 1
        sheet.cell(row=last_row + 1, column=1, value=id)
        sheet.cell(row=last_row + 1, column=2, value=models_all.loc[i, 'name'])
        sheet.cell(row=last_row + 1, column=3, value=models_all.loc[i, 'lookahead'])
        sheet.cell(row=last_row + 1, column=4, value=models_all.loc[i, 'data_start'])
        sheet.cell(row=last_row + 1, column=5, value=data_end_original)
        sheet.cell(row=last_row + 1, column=6, value=models_all.loc[i, 'learning_rate'])
        sheet.cell(row=last_row + 1, column=9, value=models_all.loc[i, 'file_name'])
        sheet.cell(row=last_row + 1, column=10, value=models_all.loc[i, 'time_frame_size'])
        sheet.cell(row=last_row + 1, column=11, value=models_all.loc[i, 'patience'])
        sheet.cell(row=last_row + 1, column=12, value=models_all.loc[i, 'min_delta'])
        sheet.cell(row=last_row + 1, column=13, value=input_size)
        sheet.cell(row=last_row + 1, column=15, value=num_seeds)
        # Add current time to the next row
        current_time = datetime.now()
        sheet.cell(row=last_row + 1, column=16, value=current_time)
        sheet.cell(row=last_row + 1, column=16).number_format = 'DD.MM.YYYY HH:MM:SS'
        sheet.cell(row=last_row + 1, column=19, value=models_all.loc[i, 'season_features_size'])
        sheet.cell(row=last_row + 1, column=20, value=models_all.loc[i, 'feature_threshold'])
        sheet.cell(row=last_row + 1, column=21, value=models_all.loc[i, 'learning_rate_scheduler'])
        sheet.cell(row=last_row + 1, column=22, value=models_all.loc[i, 'include_flights'])
        sheet.cell(row=last_row + 1, column=23, value=models_all.loc[i, 'num_roll_features'])
        sheet.cell(row=last_row + 1, column=24, value=models_all.loc[i, 'time_frame_features_size'])
        sheet.cell(row=last_row + 1, column=25, value=models_all.loc[i, 'time_frame_flights_size'])
        sheet.cell(row=last_row + 1, column=26, value=models_all.loc[i, 'average_window_size'])
        sheet.cell(row=last_row + 1, column=29, value=models_all.loc[i, 'ref_id'])
        sheet.cell(row=last_row + 1, column=27, value=mape)
        sheet.cell(row=last_row + 1, column=30, value=models_all.loc[i, 'num_epochs'])
        sheet.cell(row=last_row + 1, column=28, value=models_all.loc[i, 'average_season_features'])
        sheet.cell(row=last_row + 1, column=31, value=hmape)
        sheet.cell(row=last_row + 1, column=32, value=mape_test)
        sheet.cell(row=last_row + 1, column=33, value=hmape_test)
        sheet.cell(row=last_row + 1, column=34, value=models_all.loc[i, 'gdp_size'])
        sheet.cell(row=last_row + 1, column=35, value=models_all.loc[i, 'add_timestamp'])

        # Plot values
        if time_frame_size == 0:
            actual_values = actual_values[models_all.loc[i, 'data_start']+models_all.loc[i, 'lookahead']:].reset_index(drop=True)
        else:
            actual_values = actual_values[models_all.loc[i, 'data_start']+models_all.loc[i, 'lookahead']+time_frame_size+1:].reset_index(drop=True)
        plt.figure(figsize=(15, 8))
        plt.plot(actual_values, label='Actual Values')
        plt.plot(predicted_values, label='Predicted Values')
        plt.xlabel('Time')
        plt.ylabel('Number of Flights')
        title = 'Model: ' + models_all.loc[i, 'name']
        title += '; Feature_threshold: ' + str(models_all.loc[i, 'feature_threshold'])
        title += f';\nTime_frame_flights: {models_all.loc[i, 'time_frame_flights_size']}; Time_frame_features: {models_all.loc[i, 'time_frame_features_size']}'
        title += f'Data: {models_all.loc[i, 'file_name']};\n'
        title += f'Learning rate: {models_all.loc[i, 'learning_rate']}; Epochs: {models_all.loc[i, 'num_epochs']}'
        title += f'; MAPE: {mape:.2f}%, HMAPE: {hmape:.2f}%; MAPE test: {mape_test:.2f}%, HMAPE test: {hmape_test:.2f}%'
        plt.title(title)
        plt.legend()
        plt.grid(True)

        # Format x-axis
        # Generate date range starting from September 1997
        if time_frame_size == 0:
            start_date = pd.to_datetime('1997-09') + pd.DateOffset(months=models_all.loc[i, 'data_start'] + models_all.loc[i, 'lookahead'])
        else:
            start_date = pd.to_datetime('1997-09') + pd.DateOffset(months=models_all.loc[i, 'data_start'] + models_all.loc[i, 'lookahead'] + time_frame_size + 1)
        date_range = pd.date_range(start=start_date, periods=models_all.loc[i, 'data_end'] - models_all.loc[i, 'data_start'], freq='ME')
        # Extract years and set x-ticks for the start of each year
        x_ticks = [i for i, date in enumerate(date_range) if date.month == 1]
        x_labels = [date.year for date in date_range if date.month == 1]
        # Set x-axis
        plt.xticks(ticks=x_ticks, labels=x_labels)
        plt.xlim(0, len(predicted_values))  # Set x-axis limits

        plt.savefig(plot_path + f'{id}_sbs_values_plot.png', dpi=300)

        ids.append(id)

        # Save model
        path = sbs_path + f'{id}_model.joblib'
        if not os.path.exists(path):
            dump(models_all.loc[i, 'model'], path)
        else:
            print(f"Model file {path} already exists.")
            timestamp = datetime.now().strftime("%d-%m-%Y_%H_%M_%S")
            path = sbs_path + f'{id}_model_{timestamp}.joblib'
            dump(models_all.loc[i, 'model'], path)

        # Save history
        path = sbs_path + f'{id}_history.joblib'
        history_df = pd.DataFrame(columns=['best_seed', 'train_loss'])
        history_df['best_seed'] = models_all.loc[i, 'best_seed']
        history_df['train_loss'] = models_all.loc[i, 'train_loss']
        if not os.path.exists(path):
            dump(history_df, path)
        else:
            print(f"History file {path} already exists.")
            timestamp = datetime.now().strftime("%d-%m-%Y_%H_%M_%S")
            path = sbs_path + f'{id}_history_{timestamp}.joblib'
            dump(history_df, path)

        # Save features
        path = sbs_path + f'{id}_features_original.joblib'
        if not os.path.exists(path):
            dump(features_original, path)
        else:
            print(f"Features file {path} already exists.")
            timestamp = datetime.now().strftime("%d-%m-%Y_%H_%M_%S")
            path = sbs_path + f'{id}_features_original_{timestamp}.joblib'
            dump(features_original, path)

        # Save actual values
        path = sbs_path + f'{id}_actual_values.joblib'
        if not os.path.exists(path):
            dump(actual_values, path)
        else:
            print(f"Actual values file {path} already exists.")
            timestamp = datetime.now().strftime("%d-%m-%Y_%H_%M_%S")
            path = sbs_path + f'{id}_actual_values_{timestamp}.joblib'
            dump(actual_values, path)
        
        # Save scalers
        path = sbs_path + f'{id}_scalers.joblib'
        if not os.path.exists(path):
            dump(scalers_dict, path)
        else:
            print(f"Scalers file {path} already exists.")
            timestamp = datetime.now().strftime("%d-%m-%Y_%H_%M_%S")
            path = sbs_path + f'{id}_scalers_{timestamp}.joblib'
            dump(scalers_dict, path)
        
        # Save predicted values
        path = sbs_path + f'{id}_predicted_values.joblib'
        if not os.path.exists(path):
            dump(predicted_values, path)
        else:
            print(f"Predicted values file {path} already exists.")
            timestamp = datetime.now().strftime("%d-%m-%Y_%H_%M_%S")
            path = sbs_path + f'{id}_predicted_values_{timestamp}.joblib'
            dump(predicted_values, path)

        # Try to save the Excel file
        try:
            workbook.save(summary_path)
            rows_buffer = list()
            print("File saved successfully.")
            if i > 20:
                print("Restarting the script.")
                subprocess.Popen(["python", "start_script_sbs.py"])
                sys.exit(0)
        except PermissionError:
            workbook = openpyxl.load_workbook(summary_path)
            sheet = workbook.active
            rows_buffer.append({
                'id': id,
                'name': models_all.loc[i, 'name'],
                'lookahead': models_all.loc[i, 'lookahead'],
                'data_start': models_all.loc[i, 'data_start'],
                'data_end': data_end_original,
                'learning_rate': models_all.loc[i, 'learning_rate'],
                'file_name': models_all.loc[i, 'file_name'],
                'time_frame_size': models_all.loc[i, 'time_frame_size'],
                'patience': models_all.loc[i, 'patience'],
                'min_delta': models_all.loc[i, 'min_delta'],
                'input_size': input_size,
                'num_seeds': num_seeds,
                'less_seeds': less_seeds,
                'less_seeds_row': less_seeds_row,
                'time': current_time,
                'season_features_size': models_all.loc[i, 'season_features_size'],
                'feature_threshold': models_all.loc[i, 'feature_threshold'],
                'learning_rate_scheduler': models_all.loc[i, 'learning_rate_scheduler'],
                'include_flights': models_all.loc[i, 'include_flights'],
                'num_roll_features': models_all.loc[i, 'num_roll_features'],
                'time_frame_features_size': models_all.loc[i, 'time_frame_features_size'],
                'time_frame_flights_size': models_all.loc[i, 'time_frame_flights_size'],
                'average_window_size': models_all.loc[i, 'average_window_size'],
                'ref_id': models_all.loc[i, 'ref_id'],
                'mape': mape,
                'num_epochs': models_all.loc[i, 'num_epochs'],
                'average_season_features': models_all.loc[i, 'average_season_features'],
                'hmape': hmape,
                'mape_test': mape_test,
                'hmape_test': hmape_test,
                'gdp_size': models_all.loc[i, 'gdp_size'],
                'add_timestamp': models_all.loc[i, 'add_timestamp']
            })
            print(f"The file is open. Row with ID {id} saved in buffer.")
            print(f'Buffer size: {len(rows_buffer)}')
        
        elapsed_time = time.time() - start_time
        # Estimate remaining time
        iterations_left = total_iterations - (i + 1)
        time_per_iteration = elapsed_time / (i + 1)
        estimated_time_left = iterations_left * time_per_iteration
        # Convert elapsed time to hh:mm format
        elapsed_hours, elapsed_minutes = divmod(elapsed_time // 60, 60)
        # Convert estimated time left to hh:mm format
        estimated_hours, estimated_minutes = divmod(estimated_time_left // 60, 60)
        # Print elapsed time and estimated time left in hh:mm format
        print(f"Elapsed time: {int(elapsed_hours):02d}:{int(elapsed_minutes):02d}")
        print(f"Estimated time left: {int(estimated_hours):02d}:{int(estimated_minutes):02d}")
        # Calculate estimated time of arrival (ETA)
        eta = datetime.now() + timedelta(seconds=estimated_time_left)
        print(f"ETA: {eta.strftime('%d.%m.%Y %H:%M:%S')}\n")



    if len(rows_buffer) > 0:
        print('All models finished.')
        print('Waiting for the file to be closed...')

    while len(rows_buffer) > 0:
        time.sleep(3)
        # Add buffered rows to the sheet
        for row_data in rows_buffer:
            # Save model in summary
            if row_data['less_seeds']:
                last_row = row_data['less_seeds_row']
            else:
                # Find the last row with data in column 1
                last_row = 0
                for row in range(sheet.max_row, 0, -1):
                    if sheet.cell(row=row, column=1).value is not None:
                        last_row = row
                        break
            sheet.cell(row=last_row + 1, column=1, value=row_data['id'])
            sheet.cell(row=last_row + 1, column=2, value=row_data['name'])
            sheet.cell(row=last_row + 1, column=3, value=row_data['lookahead'])
            sheet.cell(row=last_row + 1, column=4, value=row_data['data_start'])
            sheet.cell(row=last_row + 1, column=5, value=row_data['data_end'])
            sheet.cell(row=last_row + 1, column=6, value=row_data['learning_rate'])
            sheet.cell(row=last_row + 1, column=9, value=row_data['file_name'])
            sheet.cell(row=last_row + 1, column=10, value=row_data['time_frame_size'])
            sheet.cell(row=last_row + 1, column=11, value=row_data['patience'])
            sheet.cell(row=last_row + 1, column=12, value=row_data['min_delta'])
            sheet.cell(row=last_row + 1, column=13, value=row_data['input_size'])
            sheet.cell(row=last_row + 1, column=15, value=row_data['num_seeds'])
            sheet.cell(row=last_row + 1, column=16, value=row_data['time'])
            sheet.cell(row=last_row + 1, column=16).number_format = 'DD.MM.YYYY HH:MM:SS'
            sheet.cell(row=last_row + 1, column=19, value=row_data['season_features_size'])
            sheet.cell(row=last_row + 1, column=20, value=row_data['feature_threshold'])
            sheet.cell(row=last_row + 1, column=21, value=row_data['learning_rate_scheduler'])
            sheet.cell(row=last_row + 1, column=22, value=row_data['include_flights'])
            sheet.cell(row=last_row + 1, column=23, value=row_data['num_roll_features'])
            sheet.cell(row=last_row + 1, column=24, value=row_data['time_frame_features_size'])
            sheet.cell(row=last_row + 1, column=25, value=row_data['time_frame_flights_size'])
            sheet.cell(row=last_row + 1, column=26, value=row_data['average_window_size'])
            sheet.cell(row=last_row + 1, column=29, value=row_data['ref_id'])
            sheet.cell(row=last_row + 1, column=27, value=row_data['mape'])
            sheet.cell(row=last_row + 1, column=30, value=row_data['num_epochs'])
            sheet.cell(row=last_row + 1, column=28, value=row_data['average_season_features'])
            sheet.cell(row=last_row + 1, column=31, value=row_data['hmape'])
            sheet.cell(row=last_row + 1, column=32, value=row_data['mape_test'])
            sheet.cell(row=last_row + 1, column=33, value=row_data['hmape_test'])
            sheet.cell(row=last_row + 1, column=34, value=row_data['gdp_size'])
            sheet.cell(row=last_row + 1, column=35, value=row_data['add_timestamp'])
        try:
            workbook.save(summary_path)
            rows_buffer = list()
            print("File saved successfully.")
            break
        except PermissionError:
            workbook = openpyxl.load_workbook(summary_path)
            sheet = workbook.active
            print("The file is open. Waiting for it to be closed...")

    print('All models saved in Summary.')
    return ids




if __name__ == '__main__':
    main()