# Plot forecast with min av, max av, 1 selected model
import plotly.express as px
import pandas as pd
from joblib import load
import os
import numpy as np
import plotly.offline as pyo
import plotly.graph_objs as go  
import plotly.io as pio
import openpyxl
import shutil
from datetime import datetime
from sklearn.linear_model import LinearRegression


def movingAverage(array, size, average_season_features, first=True):
            # Convert array to DataFrame if it is not already one
            if not isinstance(array, pd.DataFrame):
                idx_temp = array.index
                array = pd.DataFrame(array)
                array.index = idx_temp
            if first:
                result = np.zeros(array.shape, dtype=float)
            else:
                result = np.zeros((1, array.shape[1]), dtype=float)
            for col in range(array.shape[1]):
                if first:
                    if (not average_season_features) and (array.columns[col][:8] in ['3_months', '4_months', '6_months'] or array.columns[col][:9] in ['12_months', '17_months', '54_months']):
                            result[:, col] = array.iloc[:, col]
                    else:
                        for i in range(array.shape[0]):
                            if i < size + 1:
                                result[i, col] = np.mean(array.iloc[:i+1, col])
                            elif i >= array.shape[0] - size:
                                result[i, col] = np.mean(array.iloc[i-size:, col])
                            else:
                                result[i, col] = np.mean(array.iloc[i-size:i+1, col])
                else: # If not first
                    if not average_season_features and (array.columns[col][:8] in ['3_months', '4_months', '6_months'] or array.columns[col][:9] in ['12_months', '17_months', '54_months']):
                        result[0, col] = array.iloc[-1, col]
                    else:
                        result[0, col] = np.mean(array.iloc[-size-1:, col]) # Just last row
            if first:
                result_df = pd.DataFrame(result, columns=array.columns, index=array.index)
            else:
                result_df = pd.DataFrame(result, columns=array.columns)
            return result_df

def main():
    source_dir = '06_long_term_predictions'
    best_model = 2182
    ids = [3707,3528,3681,3478,3550,3800,3639,3674,3725,3731,3738,2182,2253,3664,3652,3736,3539,3715,3505,3748,3624,3522,3580,3657,3569]
    file_endings = '_predicted_values_2050-12.joblib'
    summary_path = 'Summary.xlsx'
    actual_values_path = r'00_raw_data\monthly_flights.csv'
    plot_name = 'forecast_plot' + '.html'
    plots_dir = '08_plots_eval'

    plot_forecast(ids, best_model, source_dir, file_endings, summary_path, actual_values_path, plot_name, plots_dir)
    return

def plot_forecast(model_ids, best_model=None, source_dir=None, file_endings=None, summary_path=None, 
                  actual_values_path=None, plot_name=None, plots_dir=None):
    
    """
    Plot long term forecast for selected models and actual values.

    Args:
        model_ids (list<int>): List of model IDs to plot.
        best_model (int): ID of the model to plot individually.
        source_dir (str): Directory containing the forecast files.
        file_endings (str): File endings of the forecast files.
        summary_path (str): Path to the summary file.
        actual_values_path (str): Path to the actual values file.
        plot_name (str): Name of the plot file.
        plots_dir (str): Directory to save the plot file.
    
    Returns:
        str: Path to the plot file.
    """
    print('Plotting forecast...')
    
    if model_ids is None or len(model_ids) == 0:
        print('No model IDs provided.')
        return
    ids = model_ids
    # Set missing parameters to default
    best_model = best_model or ids[0]
    source_dir = source_dir or '06_long_term_predictions'
    file_endings = file_endings or '_predicted_values_2050-12.joblib'
    summary_path = summary_path or 'Summary.xlsx'
    actual_values_path = actual_values_path or r'00_raw_data\monthly_flights.csv'
    plot_name = plot_name or 'forecast_plot.html'
    plots_dir = plots_dir or '08_plots_eval'

    def get_predictions(source_dir, ids, avg, conversion_rate, file_endings=file_endings):
        print('Getting predictions...')
        files = os.listdir(source_dir)  # List all files in the source directory
        preds = {}
        preds_avg = {}
        got_file = False
        print('Models missing from predictions:')
        for id in ids:
            for file in files:
                if file == f'{id}{file_endings}':
                    pred = pd.DataFrame(load(os.path.join(source_dir, file)), columns=['values'])
                    pred.index.name = 'time'
                    pred['Litres'] = pred['values'] * conversion_rate
                    preds[id] = pred
                    preds_avg[id] = movingAverage(pred, avg, True)
                    got_file = True
            if not got_file:
                print(f'{id}')
        return preds, preds_avg

    def get_add_info(ids, summary_dir=summary_path):
        print('Getting additional information...')
        workbook = openpyxl.load_workbook(summary_dir)
        sheet = workbook.active
        all_info = {}
        rows = {}
        row = None
        for id in ids:
            for r in sheet.iter_rows(min_row=1, max_row=sheet.max_row, values_only=True):
                if r[0] == id:
                    row = r
                    break
            if row is None:
                print(f'Model {id} not found in Summary.')
                continue
            info = (f'Name: {row[1]}; learning_rate: {row[5]}; time_frame_size: {row[9]}; '
                    f'input_size: {row[12]}; season_features_size: {row[18]}; include_flights: {row[21]}; '
                    f'num_roll_features: {row[22]}; \ntime_frame_features_size: {row[23]}; time_frame_flights_size: {row[24]}; '
                    f'average_window_size: {row[25]}; average_season_features: {row[27]}; num_epochs: {row[29]}; '
                    f'\nwindow_size: {int(row[8][-6:-4])}; gdp_size: {row[33]}; add_timestamp: {row[34]}')
            all_info[id] = info
            rows[id] = row
        return all_info, rows
        
    def get_actual_values():
        print('Getting actual values...')
        actual_values = pd.read_csv(actual_values_path).drop(columns=['YearMonth'])
        date_range = pd.date_range(start='1997-09', periods=len(actual_values), freq='ME')
        actual_values['time'] = date_range
        actual_values = actual_values.drop(columns=['Year', 'Month'])
        actual_values.index = actual_values['time']
        actual_values = actual_values.drop(columns=['time'])
        return actual_values
        
    def assign_pred_time(predictions, rows):
        print('Assigning prediction time...')
        for id, pred in predictions.items():
            row = rows[id]
            data_start = row[3]
            data_end = row[4]
            lookahead = row[2]
            time_frame_size = row[9]
            # Generate date range starting from September 1997
            if time_frame_size == 0:
                start_date = pd.to_datetime('1997-09') + pd.DateOffset(months=data_start + lookahead)
            else:
                start_date = pd.to_datetime('1997-09') + pd.DateOffset(months=data_start + lookahead + time_frame_size + 1)
            length = len(pred)
            date_range = pd.date_range(start=start_date, periods=length, freq='ME')
            pred.index= date_range
            pred.index.name = 'time'
            predictions[id] = pred
        return predictions

    def create_bands(fig, preds, start_date='2024-06-30', end_date='2050-12-31'):
        
        start_min_all = []
        end_min_all = []
        start_max_all = []
        end_max_all = []
        for id, pred in preds.items():
            model = LinearRegression()
            y = pred['2024-06-30':]['Litres'].values
            X = np.array(range(len(y))).reshape(-1, 1)  # Reshape to 2D array for sklearn
            model.fit(X, y)
            preds_lin = pd.DataFrame(model.predict(X), columns=['Litres'])
            preds_lin.index = pred['2024-06-30':].index
        
            diff_max = preds_lin.loc[start_date:]['Litres'].values - pred.loc[start_date:]['Litres'].values
            diff_max = diff_max.min() # -

            start_max = preds_lin.loc[start_date]['Litres'] - diff_max
            end_max = preds_lin.loc[end_date]['Litres'] - diff_max

            diff_min = preds_lin.loc[start_date:]['Litres'] - pred.loc[start_date:]['Litres']
            diff_min = diff_min.max() # +
            start_min = preds_lin.loc[start_date]['Litres'] - diff_min
            end_min = preds_lin.loc[end_date]['Litres'] - diff_min

            start_min_all.append(start_min)
            end_min_all.append(end_min)
            start_max_all.append(start_max)
            end_max_all.append(end_max)
        end_min_min = min(end_min_all)
        start_min_min = min(start_min_all)
        end_max_max = max(end_max_all)
        start_max_max = max(start_max_all)
        start_min_avg = np.mean(start_min_all)
        end_min_avg = np.mean(end_min_all)
        start_max_avg = np.mean(start_max_all)
        end_max_avg = np.mean(end_max_all)

        fig.add_trace(
            go.Scatter(
                x=[start_date, end_date],
                y=[start_max_avg, end_max_avg],
                mode='lines',
                name='max avg',
                line=dict(color='Blue', width=1.5)
            )
        )
        fig.add_trace(
            go.Scatter(
                x=[start_date, end_date],
                y=[start_min_avg, end_min_avg],
                mode='lines',
                name='min avg',
                line=dict(color='Blue', width=1.5)
            )
        )

        start_avg = np.mean([start_min_avg, start_max_avg])
        end_avg = np.mean([end_min_avg, end_max_avg])
        
        fig.add_trace(
            go.Scatter(
                x=[start_date, end_date],
                y=[start_avg, end_avg],
                mode='lines',
                name='avg',
                line=dict(color='Purple', width=2)
            )
        )
        x_points = [start_date, end_date, end_date, start_date]  # Define the x-points of the polygon
        y_points = [start_min_avg, end_min_avg, end_max_avg, start_max_avg]  # Define the y-points of the polygon

        fig.add_trace(go.Scatter(
            x=x_points + [x_points[0]],  # Close the polygon by repeating the first point
            y=y_points + [y_points[0]],  # Close the polygon by repeating the first point
            fill='toself',
            fillcolor='LightSkyBlue',
            opacity=0.5,
            line=dict(color='rgba(0,0,0,0)'),
            name='Inner'
        ))

        x_points = [start_date, end_date, end_date, start_date]  # Define the x-points of the polygon
        y_points = [start_min_min, end_min_min, end_max_max, start_max_max]
        return fig

    # Backup Summary
    # Get the base name of the source file
    base_name = os.path.basename(summary_path)
    # Get the file name and extension
    file_name, file_ext = os.path.splitext(base_name)
    # Get the current date
    current_date = datetime.now().strftime('%d.%m.%Y_%H-%M-%S')	
    # Create the new file name with date
    new_file_name = f"{file_name}_{current_date}{file_ext}"
    # Create the destination path
    dest_path = os.path.join('04_Summary_backups', new_file_name)
    # Copy the file to the new destination
    shutil.copy(summary_path, dest_path)

    # Create the initial figure
    fig = go.Figure()

    actual_values = get_actual_values()

    flights_2018 = actual_values.loc['2018-01-01':'2018-12-31', 'Flights'].sum()
    litres_2018 = 1532.63468
    conversion_rate = litres_2018 / flights_2018

    actual_values['Litres'] = actual_values['Flights'] * conversion_rate

    predictions, predictions_avg = get_predictions(source_dir, ids, 18, conversion_rate)
    


    additional_info, rows = get_add_info(ids)

    predictions = assign_pred_time(predictions, rows)
    predictions_avg = assign_pred_time(predictions_avg, rows)


    actual_values_avg = movingAverage(actual_values, 12, True)

    fig = create_bands(fig, predictions)

    pred = predictions[best_model]
    fig.add_trace(
        go.Scatter(
            x=pred['2024-06-30':].index,
            y=pred.loc['2024-06-30':,'Litres'],
            mode='markers',
            name=f'{best_model}',
            marker=dict(color='Red', size=2.5)
    ))

    # Actual values
    fig.add_trace(
                go.Scatter(
                    x=actual_values.index,
                    y=actual_values['Litres'],
                    mode='markers',
                    name='Actual values',
                    marker=dict(color='Blue', size=2.5)
                )
            )
    # Average of actual values
    fig.add_trace(
                go.Scatter(
                    x=actual_values_avg.index,
                    y=actual_values_avg['Litres'],
                    mode='lines',
                    name='Actual values avg',
                    line=dict(color='Black', width=0.5)
                )
            )

    temp_range = pd.date_range(start='1997-06', end='2051-03', freq='ME')
    fig.add_trace(
        go.Scatter(
            x=temp_range,
            y=[0] * len(temp_range),
            mode='markers',
            name='',
            marker=dict(color='White', size=1)
        )
    )
    fig.update_yaxes(range=[0, 190])  # Adjust the range as needed
    fig.update_yaxes(tickformat='~s')
    fig.update_xaxes(range=[temp_range[0], temp_range[-1]])

    file_name = f'{plots_dir}/{plot_name}'
    # Save the figure as an HTML file and open it in the browser
    pyo.plot(fig, filename=file_name, auto_open=True)
    fig.write_image(file_name + '.svg', format='svg')
    print(f'Plot saved to {file_name}.')
    return file_name

if __name__ == '__main__':
    main()