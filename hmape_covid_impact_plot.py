# SBS hmape_test vs. learning rate

import plotly.express as px
import pandas as pd
import os
import plotly.offline as pyo
import plotly.io as pio
import openpyxl
from datetime import datetime

def main():
    ids = [2365,2366,2960,2961,2962,2963,2964,2965,2966,2967,2968,2969,2970,2971,2972,2973,2974,2975,2976,2977,2978,2979,2980,2981,2982,2983,2984,2985,2986,2987,2988,2989,2990,2991,2992,2993,2994,2995,2996,2997,2998,2999,3000,3001,3002,3003,3004,3005,3006,3007,3008,3009,3010,3011,3012,3013,3014,3015,3016,3017,3018,3019,3020,3021,3022,3023,3024,3025,3026,3027,3028,3029,3030,3031,3032,3033,3034,3035,3036,3037,3038,3039,3040,3041,3042,3043,3044,3045,3046,3047,3048,3049,3050,3051,3052,3053,3054,3055,3056,3057,3058,3132,3141,3142,3143,3144,3145,3146,3147,3148,3149,3801,3802,3803,3804,3805,3806,3807,3808,3809,3810,3811,3812,3813,3814,3815,3816,3817,3818,3819,3820,3821,3822,3823,3824,3825,3826,3827,3828,3829,3830,3831,3832,3833,3834,3835,3836,3837,3838,3839,3840,3841,3842,3843,3844,3845,3846,3847,3848,3849,3850,3851,3852,3853,3854,3855,3856,3857,3858,3859,3860,3861,3862,3863,3864,3865,3866,3867,3868,3869,3870,3871,3872,3873,3874,3875,3876,3877,3878,3879,3880,3881,3882,3883,3884,3885,3886,3887,3888,3889,3890,3891,3892,3893,3894,3895,3896,3897,3898,4575,4576,4577,4578,4579,4580,4581,4582,4583,4584,4585,4586,4587,4588,4589,4590,4591,4592,4593,4594,4595,4596,4597,4598,4599,4600,4601,4602,4603]
    summary_path = 'Summary.xlsx'
    plot_name = 'sbs_covid_impact_ref_2192_epochs_1'
    plots_dir = '08_plots_eval'
    redo_plot = True
    multiply_epochs = False

    base_impact = 2.85083319448e-04

    plot_covid_impact(ids, summary_path, plot_name, plots_dir, redo_plot, multiply_epochs, base_impact)
    return

def plot_covid_impact(ids, summary_path=None, plot_name=None, plots_dir=None, redo_plot=True, multiply_epochs=False, base_impact=2.85083319448e-04, write_png=True, write_svg=True):
    """
    Plot the HMAPE vs. covid impact for the given model IDs.

    Args:
        ids (list<int>): List of model IDs to plot.
        summary_path (str): Path to the summary file.
        plot_name (str): Name of the plot.
        plots_dir (str): Directory to save the plot.
        redo_plot (bool): Whether to overwrite the existing plot.
        multiply_epochs (bool): Whether to multiply the learning rate by the number of epochs.
        base_impact (float): Base impact value.
        write_png (bool): Whether to save the plot to PNG.
        write_svg (bool): Whether to save the plot to SVG.

    Returns:
        str: Path to the saved plot.
    """
    print('Plotting HMAPE vs. covid impact...')


    # Set missing parameters do default values
    summary_path = summary_path or 'Summary.xlsx'
    plot_name = plot_name or 'sbs_covid_impact'
    plots_dir = plots_dir or '08_plots_eval'

    workbook = openpyxl.load_workbook(summary_path)
    sheet = workbook.active
    rows = {}
    row = None
    df = pd.DataFrame(columns=['id', 'hmape_test', 'learning_rate', 'num_steps', 'covid_impact'])


    for r in sheet.iter_rows(min_row=1, max_row=sheet.max_row, values_only=True):
        for id in ids:
            if r[0] == id:
                rows[id] = r
                break
    for id in ids:
        row = rows.get(id)
        if row is None:
            print(f'Model {id} not found in Summary.')
            continue
        if multiply_epochs:
            learning = row[5] * row[29]
        else:
            learning = row[5]
        covid_impact = learning / base_impact

        df_temp = pd.DataFrame({'id': [id], 'hmape_test': [row[32]], 'learning_rate': [learning], 'num_steps': [row[29]], 'covid_impact': [covid_impact]})
        df = pd.concat([df, df_temp], axis=0, ignore_index=True)

    fig = px.scatter(df, x='covid_impact', y='hmape_test', title='SBS HMAPE vs. covid impact', log_x=True)
    fig.update_traces(hovertemplate='ID: %{text}<br>Learning Rate: %{x}<br>HMAPE: %{y}<br>Num Steps: %{customdata}', 
                      text=df['id'], 
                      customdata=df['num_steps'], 
                      selector=dict(mode='markers'), marker=dict(size=5))
    fig.update_xaxes(tickformat='.2%')
    fig.update_xaxes(tickvals=[0.0001, 0.001, 0.01, 0.1, 1])

    plot_file_name = os.path.join(plots_dir, plot_name)
    if not redo_plot and os.path.exists(plot_file_name + '.html'):
        plot_file_name = f'{plot_file_name}_{datetime.now().strftime("%Y-%m-%d_%H-%M")}'

    pyo.plot(fig, filename=plot_file_name + '.html')
    if write_png:
        pio.write_image(fig, plot_file_name + '.png', format='png', width=1920, height=1080)
    if write_svg:
        fig.write_image(plot_file_name + '.svg', format='svg', scale=1.5)

    print('Plot saved.')
    return plot_file_name



if __name__ == '__main__':
    main()