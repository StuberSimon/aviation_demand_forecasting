# hmape vs input_size
import plotly.express as px
import pandas as pd
from joblib import load
import os
import numpy as np
import plotly.offline as pyo
import plotly.graph_objs as go  
import plotly.io as pio
import openpyxl
from datetime import datetime

def main():
    ids = range(2182,5000)
    summary_path = 'Summary.xlsx'
    plot_name = 'input_size_hmape'
    plots_dir = '08_plots_eval'
    redo_plot = False
    to_svg = False
    to_png = True
    print_containing_ids = False
    plot_hmape_input_size(ids, summary_path, plot_name, plots_dir, redo_plot, to_svg, to_png, print_containing_ids)
    return


def plot_hmape_input_size(ids, summary_path=None, plot_name=None, plots_dir=None, redo_plot=True, to_svg=True, to_png=True, print_containing_ids=False):

    """
    Plot the hmape vs input_size for the given model IDs.

    Args:
        ids (list<int>): List of model IDs to plot.
        summary_path (str): Path to the Summary.xlsx file.
        plot_name (str): Name of the plot.
        plots_dir (str): Directory to save the plot.
        redo_plot (bool): Whether to overwrite the existing plot.
        to_svg (bool): Whether to save the plot to SVG.
        to_png (bool): Whether to save the plot to PNG.

    Returns:
        str: Path to the saved plot.
    """
    print('Plotting hmape vs input_size...')

    if ids is None or len(ids) == 0:
        print('No model IDs provided. Exiting.')
        return
    # Set missing parameters do default values
    summary_path = summary_path or 'Summary.xlsx'
    plot_name = plot_name or 'input_size_hmape'
    plots_dir = plots_dir or '08_plots_eval'

    workbook = openpyxl.load_workbook(summary_path)
    sheet = workbook.active
    rows = {}
    row = None
    df = pd.DataFrame(columns=['id', 'hmape', 'input_size', 'time_frame_features_size', 'time_frame_flights_size'])

    def is_red(hex_color):
        # Remove the '#' character if present
        hex_color = hex_color.lstrip('#')
        
        # Convert hex to RGB
        r, g, b = int(hex_color[0:2], 16), int(hex_color[2:4], 16), int(hex_color[4:6], 16)
        
        # Define a threshold to determine if a color is "red"
        threshold = 50
        
        # Check if the red component is significantly higher than green and blue
        return (r - g > threshold) and (r - b > threshold)


    for r in sheet.iter_rows(min_row=1, max_row=sheet.max_row, values_only=True):
        for id in ids:
            if r[0] == id:
                rows[id] = r
                break
    for id in ids:
        row = rows.get(id)
        if row is None:
            print(f'Model {id} not found in Summary.')
            continue
        if row[29] > -1: # Skip SBS
            continue
        df_temp = pd.DataFrame({'id': [id], 'hmape': [row[30]], 'input_size': [row[12]], 'time_frame_features_size': [row[23]], 'time_frame_flights_size': [row[24]]})
        df = pd.concat([df, df_temp], axis=0, ignore_index=True)
        if print_containing_ids:
            print(f'{id}')

    df = df[df['input_size'] < 1500]
    df = df.sort_values(by=['input_size']).reset_index(drop=True)
    group_number = 0
    previous_input_size = df['input_size'].iloc[0] - 15  # Ensure the first value starts a new group

    # Assign groups based on the gap of at least 30
    df['group'] = 0
    for index, row in df.iterrows():
        if row['input_size'] - previous_input_size >= 15:
            group_number += 1
        df.at[index, 'group'] = group_number
        previous_input_size = row['input_size']

    colors = [color for color in px.colors.qualitative.Plotly if not is_red(color)]
    for i in range(df['group'].unique().size+1):
        df.loc[df['group'] == i, 'group_color'] = colors[i % len(colors)]

    # Plot the scatter plot with group colors
    fig = px.scatter(df, x='input_size', y='hmape', color='group_color', title='Base hmape vs. input_size', log_x=False,color_discrete_sequence=colors)
    fig.update_traces(hovertemplate='ID: %{text}<br>Input Size: %{x}<br>HMAPE: %{y}<br>Time Frame Features Size: %{customdata[0]}<br>Time Frame Flights Size: %{customdata[1]} <br>Group: %{marker.color}', 
                    text=df['id'], 
                    customdata=np.stack((df['time_frame_features_size'], df['time_frame_flights_size']), axis=-1), 
                    selector=dict(mode='markers', marker=dict(size=3)))


    # Calculate the median for each group
    group_medians = df.groupby('group')['hmape'].median().reset_index()
    group_medians.columns = ['group', 'group_median']
    group_medians['input_size'] = df.groupby('group')['input_size'].median().values

    # Merge the group medians back into the original DataFrame
    df = pd.merge(df, group_medians, on='group', how='left')

    fig.add_trace(go.Scatter(
        x=group_medians['input_size'],
        y=group_medians['group_median'],
        mode='lines+markers',
        marker=dict(color='red', size=8),
        name='Median'
    ))


    plot_file_name = os.path.join(plots_dir, plot_name)
    if not redo_plot and os.path.exists(plot_file_name + '.html'):
        plot_file_name = f'{plot_file_name}_{datetime.now().strftime("%Y-%m-%d_%H-%M")}'

    pyo.plot(fig, filename=plot_file_name + '.html')

    plot_file_name = os.path.join(plots_dir, plot_name)
    if not redo_plot and os.path.exists(plot_file_name + '.svg'):
        plot_file_name = f'{plot_file_name}_{datetime.now().strftime("%Y-%m-%d_%H-%M")}'

    if to_svg:
        pio.write_image(fig, plot_file_name + '.svg',scale=2)
    if to_png:
        pio.write_image(fig, plot_file_name + '.png',scale=2)



    selected_group = 10
    group = df[df['group'] == selected_group]
    good_models = group[group['hmape'] < 10]
    for i in range(good_models.shape[0]):
        print(f'{good_models["id"].iloc[i]}')
    
    print('Plot saved.')
    return plot_file_name



if __name__ == '__main__':
    main()