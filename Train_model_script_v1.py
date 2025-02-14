class DummyFile(object):
    def write(self, x): pass
    def flush(self): pass

import os
# Suppress TensorFlow warnings
os.environ['TF_CPP_MIN_LOG_LEVEL'] = '3'
import numpy as np
import pandas as pd
from joblib import dump, load
import tensorflow as tf
from tensorflow import keras
from keras import layers
from tqdm import tqdm
import random
from datetime import datetime, timedelta
from tsfresh.utilities.dataframe_functions import roll_time_series
from sklearn.preprocessing import MinMaxScaler
import openpyxl
import time
import contextlib
import shutil
import gc
import subprocess
dummy_file = DummyFile()


with contextlib.redirect_stdout(dummy_file), contextlib.redirect_stderr(dummy_file):
    def set_random_seed(seed):
        np.random.seed(seed)
        tf.random.set_seed(seed)
        random.seed(seed)
        os.environ['PYTHONHASHSEED'] = str(seed)
        os.environ['TF_DETERMINISTIC_OPS'] = '1'
    # Define a learning rate scheduler function
    def scheduler(epoch, lr):
        if epoch < 20:
            return float(lr)
        else:
            return float(lr * tf.math.exp(-0.1))
def movingAverage(array, size, average_season_features):
    # Convert array to DataFrame if it is not already one
    if not isinstance(array, pd.DataFrame):
        idx_temp = array.index
        array = pd.DataFrame(array)
        array.index = idx_temp
    result = np.zeros(array.shape, dtype=float)
    for col in range(array.shape[1]):
        if (not average_season_features) and (array.columns[col][:8] in ['3_months', '4_months', '6_months'] or array.columns[col][:9] in ['12_months', '17_months', '54_months']):
            result[:, col] = array.iloc[:, col]
        else:
            for i in range(array.shape[0]):
                if i < size:
                    result[i, col] = np.mean(array.iloc[:i+1, col])
                elif i >= array.shape[0] - size:
                    result[i, col] = np.mean(array.iloc[i-size:, col])
                else:
                    result[i, col] = np.mean(array.iloc[i-size:i+1, col])
    result_df = pd.DataFrame(result, columns=array.columns, index=array.index)
    return result_df


def main():   
    # Base path for the extracted features with '/'
    features_path = '01_extracted_features/'
    models_path = '02_models/'
    summary_path = 'Summary.xlsx'
    # File names with .csv extension
    feature_file_names = ['all_features_window-size=24.csv']#  
    # Shift the labels by this amount
    lookahead_list = [1] # Default 1
    data_start_list = [0] # Default 0
    data_end_list = [269] # Default 269
    learning_rates_list = [1e-5] # Default 1e-5
    patience_list = [300] # Default 300
    min_delta_list = [1e-7] # Default 1e-7
    add_season_features_list = [True] # Default True
    feature_threshold_list = [0.9] # Default 0.9
    learning_rate_scheduler_list = [True] # Default True
    include_flights_list = [True] # Default True
    num_roll_features_list = [32] # Default 32
    time_frame_features_size_list = [24] # Default 24
    time_frame_flights_size_list = [24] # Default 24
    average_window_size_list = [0]              # Default 0
    average_season_features_list = [False]       # Default False
    add_gdp_list = [True] # Default True
    add_timestamp_list = [1] # Default 1

    dump_train = False # Default False
    restart_script = False # Default False
    num_seeds = 6 # Default 6
    seed_start = 42 # Default 42
    max_epochs = 3000 # Default 3000
    season_features = load(features_path + 'season_features_3_4_6_12_17_54.joblib')
    feature_importance = load(features_path + '1290_feature_importance_df.joblib')
    gdp_features = load(features_path + 'gdp_features_v1.joblib').reset_index(drop=True)

    # Models
    with contextlib.redirect_stdout(dummy_file), contextlib.redirect_stderr(dummy_file):
        def create_model1(size):
            return keras.Sequential([
                layers.Dense(512, activation='relu'),
                layers.Dense(256, activation='relu'),
                layers.Dense(128, activation='relu'),
                layers.Dense(64, activation='relu'),
                layers.Dense(32, activation='relu'),
                layers.Dense(16, activation='relu'),
                layers.Dense(8, activation='relu'),
                layers.Dense(4, activation='relu'),
                layers.Dense(2, activation='relu'),
                layers.Dense(1)  # Output layer with 1 unit
                ])
        def create_model2(size):
            return keras.Sequential([
                layers.Dense(1000, activation='relu'),
                layers.Dense(500, activation='relu'),
                layers.Dense(1)  # Output layer with 1 unit
                ])
        def create_model3(size):
            return keras.Sequential([
                layers.Dense(1024, activation='relu'),
                layers.Dense(512, activation='relu'),
                layers.Dense(256, activation='relu'),
                layers.Dense(128, activation='relu'),
                layers.Dense(64, activation='relu'),
                layers.Dense(32, activation='relu'),
                layers.Dense(16, activation='relu'),
                layers.Dense(8, activation='relu'),
                layers.Dense(4, activation='relu'),
                layers.Dense(2, activation='relu'),
                layers.Dense(1)  # Output layer with 1 unit
            ])
        def create_model4(size):
            return keras.Sequential([
                layers.Dense(2000, activation='relu'),
                layers.Dense(2000, activation='relu'),
                layers.Dense(2000, activation='relu'),
                layers.Dense(2000, activation='relu'),
                layers.Dense(2000, activation='relu'),
                layers.Dense(2000, activation='relu'),
                layers.Dense(2000, activation='relu'),
                layers.Dense(2000, activation='relu'),
                layers.Dense(2000, activation='relu'),
                layers.Dense(2000, activation='relu'),
                layers.Dense(2000, activation='relu'),
                layers.Dense(2000, activation='relu'),
                layers.Dense(1000, activation='relu'),
                layers.Dense(500, activation='relu'),
                layers.Dense(250, activation='relu'),
                layers.Dense(125, activation='relu'),
                layers.Dense(62, activation='relu'),
                layers.Dense(31, activation='relu'),
                layers.Dense(15, activation='relu'),
                layers.Dense(7, activation='relu'),
                layers.Dense(3, activation='relu'),
                layers.Dense(1)  # Output layer with 1 unit
                ])
        def create_model5(size):
            return keras.Sequential([
                layers.Dense(1000, activation='relu'),
                layers.Dense(1000, activation='relu'),
                layers.Dense(1000, activation='relu'),
                layers.Dense(1000, activation='relu'),
                layers.Dense(1000, activation='relu'),
                layers.Dense(1000, activation='relu'),
                layers.Dense(1000, activation='relu'),
                layers.Dense(1000, activation='relu'),
                layers.Dense(1000, activation='relu'),
                layers.Dense(1000, activation='relu'),
                layers.Dense(1000, activation='relu'),
                layers.Dense(1000, activation='relu'),
                layers.Dense(500, activation='relu'),
                layers.Dense(250, activation='relu'),
                layers.Dense(125, activation='relu'),
                layers.Dense(62, activation='relu'),
                layers.Dense(31, activation='relu'),
                layers.Dense(1)  # Output layer with 1 unit
                ])

    model_functions = [create_model3]
    model_names = ['Relu 1024 512 256 128 64 32 16 8 4 2 1']
    prioritised_models = []
    
    ids = train_models(feature_file_names, data_start_list, data_end_list, learning_rates_list, patience_list, min_delta_list, add_season_features_list,
                        feature_threshold_list, learning_rate_scheduler_list, include_flights_list, num_roll_features_list, time_frame_features_size_list,
                        time_frame_flights_size_list, average_window_size_list, average_season_features_list, add_gdp_list, add_timestamp_list, dump_train, num_seeds,
                        seed_start, max_epochs, season_features,feature_importance, gdp_features, model_functions, model_names, summary_path,
                        features_path, models_path, restart_script)

    return

def train_models(feature_file_names=None, data_start_list=None, data_end_list=None, learning_rates_list=None, patience_list=None, min_delta_list=None, add_season_features_list=None,
                 feature_threshold_list=None, learning_rate_scheduler_list=None, include_flights_list=None, num_roll_features_list=None, time_frame_features_size_list=None,
                 time_frame_flights_size_list=None, average_window_size_list=None, average_season_features_list=None, add_gdp_list=None, add_timestamp_list=None, dump_train=False, num_seeds=None,
                 seed_start=None, max_epochs=None, season_features=None, feature_importance=None, gdp_features=None, model_functions=None, model_names=None, summary_path=None,
                 features_path=None, models_path=None, restart_script=False, features_list=None):
    """
    Trains models with different parameters and saves them to disk.

    Args:
        feature_file_names (list<string>): List of file names with the extracted features.
        data_start_list (list<int>): List of start indices for the training data.
        data_end_list (list<int>): List of end indices for the training data.
        learning_rates_list (list<int>): List of learning rates for the models.
        patience_list (list<int>): List of patience values for the early stopping callback.
        min_delta_list (list<double>): List of minimum delta values for the early stopping callback.
        add_season_features_list (list<bool>): List of boolean values to add season features to the models.
        feature_threshold_list (list<double>): List of feature threshold values for the feature selection.
        learning_rate_scheduler_list (list<bool>): List of boolean values to use a learning rate scheduler.
        include_flights_list (list<bool>): List of boolean values to include flight data directly in the models.
        num_roll_features_list (list<int>): List of number of rolling features to add to the models.
        time_frame_features_size_list (list<int>): List of time frame sizes for the feature rolling.
        time_frame_flights_size_list (list<int>): List of time frame sizes for the flight rolling.
        average_window_size_list (list<int>): List of window sizes for the moving average.
        average_season_features_list (list<bool>): List of boolean values to average season features.
        add_gdp_list (list<bool>): List of boolean values to add GDP features to the models.
        add_timestamp_list (list<bool>): List of boolean values to add a timestamp feature to the models.
        dump_train (bool): Boolean value to save the training data to disc.
        num_seeds (int): Number of seeds to use for training.
        seed_start (int): Starting seed for the random number generator.
        max_epochs (int): Maximum number of epochs for training.
        season_features (DataFrame): DataFrame with season features.
        feature_importance (DataFrame): DataFrame with feature importance values.
        gdp_features (DataFrame): DataFrame with GDP features.
        model_functions (list<function>): List of model functions to use for training.
        model_names (list<string>): List of model names.
        summary_path (string): Path to the summary excel file.
        features_path (string): Path to the extracted features.
        models_path (string): Path to save the trained models.
        restart_script (bool): Boolean value to restart the script after 30 models. Set to true if script hogs memory.
        features_list (list<DataFrame>): List of DataFrames with the extracted features.

    Returns:
        list<int>: List of new model IDs.
    """

    # Set missing parameters to default
    def create_model_default(size):
            return keras.Sequential([
                layers.Dense(1024, activation='relu'),
                layers.Dense(512, activation='relu'),
                layers.Dense(256, activation='relu'),
                layers.Dense(128, activation='relu'),
                layers.Dense(64, activation='relu'),
                layers.Dense(32, activation='relu'),
                layers.Dense(16, activation='relu'),
                layers.Dense(8, activation='relu'),
                layers.Dense(4, activation='relu'),
                layers.Dense(2, activation='relu'),
                layers.Dense(1)  # Output layer with 1 unit
            ])
    feature_file_names = feature_file_names or ['all_features_window-size=24.csv']
    lookahead_list = [1] # Changing lookahead is depreciated. Values > 1 might not work with feature rolling.
    data_start_list = data_start_list or [0]
    data_end_list = data_end_list or [269]
    learning_rates_list = learning_rates_list or [1e-5]
    patience_list = patience_list or [300]
    min_delta_list = min_delta_list or [1e-7]
    add_season_features_list = add_season_features_list or [True]
    feature_threshold_list = feature_threshold_list or [0.9]
    learning_rate_scheduler_list = learning_rate_scheduler_list or [True]
    include_flights_list = include_flights_list or [True]
    num_roll_features_list = num_roll_features_list or [32]
    time_frame_features_size_list = time_frame_features_size_list or [24]
    time_frame_flights_size_list = time_frame_flights_size_list or [24]
    average_window_size_list = average_window_size_list or [0]
    average_season_features_list = average_season_features_list or [False]
    add_gdp_list = add_gdp_list or [True]
    add_timestamp_list = add_timestamp_list or [1]
    model_functions = model_functions or [create_model_default]
    model_names = model_names or ['Relu 1024 512 256 128 64 32 16 8 4 2 1']
    prioritised_models = [] # Prioritising models is depreciated. It can only be used to move models to the top of the list after restarting the script.
    summary_path = summary_path or 'Summary.xlsx'
    features_path = features_path or '01_extracted_features/'
    models_path = models_path or '02_models/'
    num_seeds = num_seeds or 6
    seed_start = seed_start or 42
    max_epochs = max_epochs or 3000
    season_features = season_features or load(features_path + 'season_features_3_4_6_12_17_54.joblib')
    feature_importance = feature_importance or load(features_path + '1290_feature_importance_df.joblib')
    gdp_features = gdp_features or load(features_path + 'gdp_features_v1.joblib').reset_index(drop=True)

    # Backup Summary
    # Get the base name of the source file
    base_name = os.path.basename(summary_path)
    # Get the file name and extension
    file_name, file_ext = os.path.splitext(base_name)
    # Get the current date
    current_date = datetime.now().strftime('%d.%m.%Y_%H-%M-%S')	
    # Create the new file name with date
    new_file_name = f"{file_name}_{current_date}{file_ext}"
    # Create the destination path
    dest_path = os.path.join('04_Summary_backups', new_file_name)
    # Copy the file to the new destination
    shutil.copy(summary_path, dest_path)

    features_raw_list = list()
    if features_list is None:
        for file_name in feature_file_names:
            data = pd.read_csv(features_path + file_name).drop(['YearMonth'], axis=1)
            features_raw_list.append(data)
    else:
        features_raw_list = features_list
    # Build df for all models
    models_all = pd.DataFrame(
        columns=['name', 'lookahead', 'data_start', 'data_end', 'learning_rate', 'train_loss',
                'val_loss', 'file_name', 'model', 'data_idx', 'best_seed', 'model_func',
                'time_frame_size', 'patience', 'min_delta', 'input_size', 'add_season_features',
                'season_features_size', 'feature_threshold', 'learning_rate_scheduler', 'include_flights',
                'num_roll_features', 'time_frame_features_size', 'time_frame_flights_size', 'average_window_size',
                'average_season_features', 'skip_me', 'add_gdp', 'gdp_size', 'add_timestamp'],
        index=range(len(lookahead_list) * len(data_start_list) * len(data_end_list)
                    * len(feature_file_names) * len(model_names) * len(learning_rates_list)
                    * len(patience_list) * len(min_delta_list) * len(add_season_features_list)
                    * len(feature_threshold_list) * len(learning_rate_scheduler_list) * len(include_flights_list)
                    * len(num_roll_features_list) * len(time_frame_features_size_list) * len(time_frame_flights_size_list)
                    * len(average_window_size_list) * len(average_season_features_list) * len(add_gdp_list) * len(add_timestamp_list))
    )

    for i in range(len(models_all)):
        models_all.loc[i, 'lookahead'] = lookahead_list[i % len(lookahead_list)]
        models_all.loc[i, 'data_start'] = data_start_list[(i // len(lookahead_list)) % len(data_start_list)]
        models_all.loc[i, 'data_end'] = data_end_list[(i // (len(lookahead_list) * len(data_start_list))) % len(data_end_list)]
        models_all.loc[i, 'file_name'] = feature_file_names[(i // (len(lookahead_list) * len(data_start_list) * len(data_end_list))) % len(feature_file_names)]
        models_all.loc[i, 'name'] = model_names[(i // (len(lookahead_list) * len(data_start_list) * len(data_end_list) * len(feature_file_names))) % len(model_names)]
        models_all.loc[i, 'data_idx'] = (i // (len(lookahead_list) * len(data_start_list) * len(data_end_list))) % len(feature_file_names)
        models_all.loc[i, 'learning_rate'] = learning_rates_list[(i // (len(lookahead_list) * len(data_start_list) * len(data_end_list) * len(feature_file_names) * len(model_names))) % len(learning_rates_list)]
        models_all.loc[i, 'model_func'] = model_functions[(i // (len(lookahead_list) * len(data_start_list) * len(data_end_list) * len(feature_file_names))) % len(model_names)]
        models_all.loc[i, 'patience'] = patience_list[(i // (len(lookahead_list) * len(data_start_list) * len(data_end_list) * len(feature_file_names) * len(model_names) * len(learning_rates_list))) % len(patience_list)]
        models_all.loc[i, 'min_delta'] = min_delta_list[(i // (len(lookahead_list) * len(data_start_list) * len(data_end_list) * len(feature_file_names) * len(model_names) * len(learning_rates_list) * len(patience_list))) % len(min_delta_list)]
        models_all.loc[i, 'add_season_features'] = add_season_features_list[(i // (len(lookahead_list) * len(data_start_list) * len(data_end_list) * len(feature_file_names) * len(model_names) * len(learning_rates_list) * len(patience_list) * len(min_delta_list))) % len(add_season_features_list)]
        models_all.loc[i, 'feature_threshold'] = feature_threshold_list[(i // (len(lookahead_list) * len(data_start_list) * len(data_end_list) * len(feature_file_names) * len(model_names) * len(learning_rates_list) * len(patience_list) * len(min_delta_list) * len(add_season_features_list))) % len(feature_threshold_list)]
        models_all.loc[i, 'learning_rate_scheduler'] = learning_rate_scheduler_list[(i // (len(lookahead_list) * len(data_start_list) * len(data_end_list) * len(feature_file_names) * len(model_names) * len(learning_rates_list) * len(patience_list) * len(min_delta_list) * len(add_season_features_list) * len(feature_threshold_list))) % len(learning_rate_scheduler_list)]
        models_all.loc[i, 'include_flights'] = include_flights_list[(i // (len(lookahead_list) * len(data_start_list) * len(data_end_list) * len(feature_file_names) * len(model_names) * len(learning_rates_list) * len(patience_list) * len(min_delta_list) * len(add_season_features_list) * len(feature_threshold_list) * len(learning_rate_scheduler_list))) % len(include_flights_list)]
        models_all.loc[i, 'num_roll_features'] = num_roll_features_list[(i // (len(lookahead_list) * len(data_start_list) * len(data_end_list) * len(feature_file_names) * len(model_names) * len(learning_rates_list) * len(patience_list) * len(min_delta_list) * len(add_season_features_list) * len(feature_threshold_list) * len(learning_rate_scheduler_list) * len(include_flights_list))) % len(num_roll_features_list)]
        models_all.loc[i, 'time_frame_features_size'] = time_frame_features_size_list[(i // (len(lookahead_list) * len(data_start_list) * len(data_end_list) * len(feature_file_names) * len(model_names) * len(learning_rates_list) * len(patience_list) * len(min_delta_list) * len(add_season_features_list) * len(feature_threshold_list) * len(learning_rate_scheduler_list) * len(include_flights_list) * len(num_roll_features_list))) % len(time_frame_features_size_list)]
        models_all.loc[i, 'time_frame_flights_size'] = time_frame_flights_size_list[(i // (len(lookahead_list) * len(data_start_list) * len(data_end_list) * len(feature_file_names) * len(model_names) * len(learning_rates_list) * len(patience_list) * len(min_delta_list) * len(add_season_features_list) * len(feature_threshold_list) * len(learning_rate_scheduler_list) * len(include_flights_list) * len(num_roll_features_list) * len(time_frame_features_size_list))) % len(time_frame_flights_size_list)]
        models_all.loc[i, 'average_window_size'] = average_window_size_list[(i // (len(lookahead_list) * len(data_start_list) * len(data_end_list) * len(feature_file_names) * len(model_names) * len(learning_rates_list) * len(patience_list) * len(min_delta_list) * len(add_season_features_list) * len(feature_threshold_list) * len(learning_rate_scheduler_list) * len(include_flights_list) * len(num_roll_features_list) * len(time_frame_features_size_list) * len(time_frame_flights_size_list))) % len(average_window_size_list)]
        models_all.loc[i, 'average_season_features'] = average_season_features_list[(i // (len(lookahead_list) * len(data_start_list) * len(data_end_list) * len(feature_file_names) * len(model_names) * len(learning_rates_list) * len(patience_list) * len(min_delta_list) * len(add_season_features_list) * len(feature_threshold_list) * len(learning_rate_scheduler_list) * len(include_flights_list) * len(num_roll_features_list) * len(time_frame_features_size_list) * len(time_frame_flights_size_list) * len(average_window_size_list))) % len(average_season_features_list)]
        models_all.loc[i, 'add_gdp'] = add_gdp_list[(i // (len(lookahead_list) * len(data_start_list) * len(data_end_list) * len(feature_file_names) * len(model_names) * len(learning_rates_list) * len(patience_list) * len(min_delta_list) * len(add_season_features_list) * len(feature_threshold_list) * len(learning_rate_scheduler_list) * len(include_flights_list) * len(num_roll_features_list) * len(time_frame_features_size_list) * len(time_frame_flights_size_list) * len(average_window_size_list) * len(average_season_features_list))) % len(add_gdp_list)]
        models_all.loc[i, 'add_timestamp'] = add_timestamp_list[(i // (len(lookahead_list) * len(data_start_list) * len(data_end_list) * len(feature_file_names) * len(model_names) * len(learning_rates_list) * len(patience_list) * len(min_delta_list) * len(add_season_features_list) * len(feature_threshold_list) * len(learning_rate_scheduler_list) * len(include_flights_list) * len(num_roll_features_list) * len(time_frame_features_size_list) * len(time_frame_flights_size_list) * len(average_window_size_list) * len(average_season_features_list) * len(add_gdp_list))) % len(add_timestamp_list)]

    # Shuffle models_all
    models_all = models_all.sample(frac=1, random_state=seed_start).reset_index(drop=True)

    # Save models_all in csv
    models_all.to_excel('models_all.xlsx', index=True)
    # Move prioritized models to the top of the list
    prioritized_mask = models_all.index.isin(prioritised_models)
    models_all = pd.concat([models_all[prioritized_mask], models_all[~prioritized_mask]]).reset_index(drop=True)
    models_all.to_excel('models_all_prio.xlsx', index=True)
    train_dataset_list = list()
    test_dataset_list = list()
    idx_list = list()
    features_original_list = list()
    workbook = openpyxl.load_workbook(summary_path)
    sheet = workbook.active
    rows_buffer = list()
    ids = list()
    for i in range(len(models_all)):
        scalers_dict = {}
        features_raw = features_raw_list[models_all.loc[i, 'data_idx']]

        important_features = feature_importance[feature_importance['cumulative_importance'] <= models_all.loc[i, 'feature_threshold']]
        important_features = important_features.sort_values(by='cumulative_importance', ascending=True)
        valid_features = [feature for feature in important_features['feature'] if feature in features_raw.columns]
        valid_features.append('Flights')
        features_raw = features_raw[valid_features]
        

        if models_all.loc[i, 'add_season_features']:
            models_all.loc[i, 'season_features_size'] = season_features.shape[1]
            features_raw = pd.concat([features_raw, season_features], axis=1)
        else:
            models_all.loc[i, 'season_features_size'] = 0    
        if models_all.loc[i, 'add_gdp']:
            models_all.loc[i, 'gdp_size'] = gdp_features.shape[1]
            features_raw = pd.concat([gdp_features.iloc[:,:models_all.loc[i,'gdp_size']], features_raw], axis=1)
        else:
            models_all.loc[i, 'gdp_size'] = 0
        if models_all.loc[i, 'add_timestamp'] == 1:
            features_raw['Timestamp'] = features_raw.index


        lookahead = models_all.loc[i, 'lookahead']
        num_roll_features = models_all.loc[i, 'num_roll_features']
        time_frame_features_size = models_all.loc[i, 'time_frame_features_size'] - 1
        time_frame_flights_size = models_all.loc[i, 'time_frame_flights_size'] - 1
        features_raw = features_raw.iloc[models_all.loc[i, 'data_start']:models_all.loc[i, 'data_end']]
        if time_frame_flights_size == 0 and time_frame_features_size == 0:
            input_size = features_raw.shape[1]
        else:
            input_size = features_raw.shape[1] + num_roll_features * (time_frame_features_size) + time_frame_flights_size

        models_all.loc[i, 'input_size'] = input_size

        # Check if model already in Summary
        model_in_summary = False
        less_seeds_row = 0
        less_seeds_best = seed_start
        less_seeds = False
        time_frame_size = max(models_all.loc[i, 'time_frame_features_size'], models_all.loc[i, 'time_frame_flights_size']) - 1
        models_all.loc[i, 'time_frame_size'] = time_frame_size

        for row in sheet.iter_rows(min_row=1, values_only=True):
            model_in_summary = row[1] == models_all.loc[i, 'name'] and row[2] == models_all.loc[i, 'lookahead'] and row[3] == models_all.loc[i, 'data_start']
            model_in_summary = model_in_summary and row[4] == models_all.loc[i, 'data_end'] and row[5] == models_all.loc[i, 'learning_rate'] 
            model_in_summary = model_in_summary and row[8] == models_all.loc[i, 'file_name'] and row[9] == models_all.loc[i, 'time_frame_size']
            model_in_summary = model_in_summary and row[10] == models_all.loc[i, 'patience'] and row[11] == models_all.loc[i, 'min_delta']
            model_in_summary = model_in_summary and row[12] == input_size
            model_in_summary = model_in_summary and row[18] == models_all.loc[i, 'season_features_size']
            model_in_summary = model_in_summary and row[19] == models_all.loc[i, 'feature_threshold']
            model_in_summary = model_in_summary and row[20] == models_all.loc[i, 'learning_rate_scheduler']
            model_in_summary = model_in_summary and row[21] == models_all.loc[i, 'include_flights']
            model_in_summary = model_in_summary and row[22] == models_all.loc[i, 'num_roll_features']
            model_in_summary = model_in_summary and row[23] == models_all.loc[i, 'time_frame_features_size']
            model_in_summary = model_in_summary and row[24] == models_all.loc[i, 'time_frame_flights_size']
            model_in_summary = model_in_summary and row[25] == models_all.loc[i, 'average_window_size']
            model_in_summary = model_in_summary and row[27] == models_all.loc[i, 'average_season_features']
            model_in_summary = model_in_summary and row[28] == -1 and row[29] == -1
            model_in_summary = model_in_summary and row[33] == models_all.loc[i, 'gdp_size']
            model_in_summary = model_in_summary and row[34] == models_all.loc[i, 'add_timestamp']

            if model_in_summary and row[14] is not None and not isinstance(row[14], str) and row[14] < num_seeds:
                model_in_summary = False            
            if model_in_summary:
                print(f'Model {i} already in Summary with ID {row[0]}.')
                break
        models_all.loc[i, 'skip_me'] = model_in_summary
        if model_in_summary:
            continue    

        
        train_labels = features_raw['Flights'].shift(-lookahead).dropna()
        features_original = features_raw.copy()
        features_original['labels'] = features_raw['Flights']
        if lookahead > 0:
            features_raw = features_raw.iloc[:-lookahead]
            features_raw['labels'] = train_labels
            features_raw.reset_index(drop=True, inplace=True)
        if time_frame_size == 0:
            train_dataset = features_raw.sample(frac=0.9, random_state=seed_start)
            test_dataset = features_raw.drop(train_dataset.index)
        else:
            train_dataset = features_raw.copy()
            test_dataset = features_raw.drop(train_dataset.index)
        
        idx = np.concatenate([train_dataset.index, test_dataset.index])
        train_dataset_list.append(train_dataset)
        test_dataset_list.append(test_dataset)
        idx_list.append(idx)
        features_original_list.append(features_original)

    features_raw_list = list()
    gc.collect()
    print(f'Models finished: {sum(models_all["skip_me"])}/{len(models_all)}.')
    models_all = models_all[models_all['skip_me'] == False].reset_index(drop=True)

    # Split dataset into features and labels
    start_time = time.time()
    total_iterations = len(models_all)
    for i in range(len(models_all)):
        input_size = models_all.loc[i, 'input_size']
        train_dataset = train_dataset_list[i]
        test_dataset = test_dataset_list[i]
        idx = idx_list[i]
        features_original = features_original_list[i]
        # Normalize the data
        for column in train_dataset.columns:
            scaler = MinMaxScaler()
            train_dataset[column] = scaler.fit_transform(train_dataset[column].values.reshape(-1, 1))
            if test_dataset.shape[0] > 0:
                test_dataset[column] = scaler.transform(test_dataset[column].values.reshape(-1, 1))
            features_original[column] = scaler.transform(features_original[column].values.reshape(-1, 1))
            scalers_dict[column] = scaler
        if models_all.loc[i, 'include_flights']:
            features_original = features_original.drop(['labels'], axis=1)
        else:
            features_original = features_original.drop(['labels', 'Flights'], axis=1)

        lookahead = models_all.loc[i, 'lookahead']
        time_frame_features_size = models_all.loc[i, 'time_frame_features_size'] - 1
        time_frame_flights_size = models_all.loc[i, 'time_frame_flights_size'] - 1
        time_frame_size = models_all.loc[i, 'time_frame_size']
        num_roll_features = models_all.loc[i, 'num_roll_features']
        
        if time_frame_size > 0 and train_dataset.shape[0] >= time_frame_size:
            # Select first ten columns from train_dataset and append 'Flights' if not already in
            if time_frame_features_size > 0:
                selected_columns = train_dataset.copy().drop(['labels', 'Flights'], axis=1).columns[:num_roll_features].tolist()
            else:
                selected_columns = []
            # Train features
            if models_all.loc[i, 'include_flights'] and time_frame_flights_size == 0:
                train_features = train_dataset.copy().drop(columns=selected_columns).drop(['labels'], axis=1)[time_frame_size:].reset_index(drop=True)
            elif not models_all.loc[i, 'include_flights']:
                train_features = train_dataset.copy().drop(columns=selected_columns).drop(['labels', 'Flights'], axis=1)[time_frame_size:].reset_index(drop=True)
            else:
                selected_columns.append('Flights')
                train_features = train_dataset.copy().drop(columns=selected_columns, axis=1).drop(['labels'], axis=1)[time_frame_size:].reset_index(drop=True)
            with tqdm(total=len(selected_columns), desc=f'Rolling train for model {i}/{len(models_all)-1}') as pbar_inner:
                for feature in selected_columns:
                    train_dataset_melted = train_dataset[[feature]].copy()
                    train_dataset_melted['Timestamp'] = train_dataset.index
                    train_dataset_melted['Symbols'] = 'A'
                    # Roll time series
                    with contextlib.redirect_stdout(dummy_file), contextlib.redirect_stderr(dummy_file):
                        if feature == 'Flights' and time_frame_flights_size > 0 and models_all.loc[i, 'include_flights']:
                            train_features_rolled = roll_time_series(train_dataset_melted, column_id='Symbols',
                                                                    column_sort='Timestamp', max_timeshift=time_frame_flights_size,
                                                                    min_timeshift=time_frame_flights_size)
                        elif feature != 'Flights' and time_frame_features_size > 0:
                            train_features_rolled = roll_time_series(train_dataset_melted, column_id='Symbols',
                                                                    column_sort='Timestamp', max_timeshift=time_frame_features_size,
                                                                    min_timeshift=time_frame_features_size)
                        else:
                            continue
                    # Transform rolled features
                    grouped = train_features_rolled.groupby('id')
                    dfs = {name: group for name, group in grouped}
                    train_features_temp = pd.DataFrame()
                    for key in dfs.keys():
                        df = dfs[key].drop(['id', 'Symbols'], axis=1)
                        df.index = df['Timestamp'].values
                        df = df.drop(['Timestamp'], axis=1)
                        idx_temp = df.index[0]
                        # Transpose the DataFrame
                        df_reshaped = df.T
                        # Reset the index to ensure the original index is preserved
                        df_reshaped.columns = [f'{feature}_{i}' for i in range(0, len(df_reshaped.columns))]
                        df_reshaped.index = [idx_temp]
                        train_features_temp = pd.concat([train_features_temp, df_reshaped], axis=0)
                    if feature == 'Flights':
                        if time_frame_flights_size < time_frame_features_size:
                            train_features_temp = train_features_temp[time_frame_features_size-time_frame_flights_size:].reset_index(drop=True)
                    else:
                        if time_frame_features_size < time_frame_flights_size:
                            train_features_temp = train_features_temp[time_frame_flights_size-time_frame_features_size:].reset_index(drop=True)
                    train_features = pd.concat([train_features, train_features_temp], axis=1)
                    pbar_inner.update(1)
            
            # Test features    
            test_features = pd.DataFrame()
            
            if test_dataset.shape[0] >= time_frame_size:
                selected_columns = test_dataset.copy().drop(['labels', 'Flights'], axis=1).columns[:num_roll_features].tolist()
                if models_all.loc[i, 'include_flights'] and time_frame_flights_size == 0:
                    test_features = test_dataset.copy().drop(columns=selected_columns).drop(['labels'], axis=1)
                elif not models_all.loc[i, 'include_flights']:
                    test_features = test_dataset.copy().drop(columns=selected_columns).drop(['labels', 'Flights'], axis=1)
                else:
                    selected_columns.append('Flights')
                    test_features = test_dataset.copy().drop(columns=selected_columns).drop(['labels'], axis=1)

                for feature in tqdm(selected_columns, desc=f'Rolling test for model {i+1}/{len(models_all)}'):
                    test_dataset_melted = test_dataset[[feature]].copy()
                    test_dataset_melted['Timestamp'] = test_dataset.index
                    test_dataset_melted['Symbols'] = 'A'
                    # Roll time series
                    with contextlib.redirect_stdout(dummy_file), contextlib.redirect_stderr(dummy_file):
                        if feature == 'Flights' and time_frame_flights_size > 0 and models_all.loc[i, 'include_flights']:
                            test_features_rolled = roll_time_series(test_dataset_melted, column_id='Symbols',
                                                                    column_sort='Timestamp', max_timeshift=time_frame_flights_size,
                                                                    min_timeshift=time_frame_flights_size)
                        elif feature != 'Flights' and time_frame_features_size > 0:
                            test_features_rolled = roll_time_series(test_dataset_melted, column_id='Symbols',
                                                                    column_sort='Timestamp', max_timeshift=time_frame_features_size,
                                                                    min_timeshift=time_frame_features_size)
                        else:
                            continue
                    # Transform rolled features
                    grouped = test_features_rolled.groupby('id')
                    dfs = {name: group for name, group in grouped}
                    test_features_temp = pd.DataFrame()
                    first = True
                    for key in dfs.keys():
                        if first:
                            first = False
                            continue
                        df = dfs[key].drop(['id', 'Symbols'], axis=1)
                        df.index = df['Timestamp'].values
                        df = df.drop(['Timestamp'], axis=1)
                        idx_temp = df.index[0]
                        # Transpose the DataFrame
                        df_reshaped = df.T
                        # Reset the index to ensure the original index is preserved
                        df_reshaped.columns = [f'{feature}_{i}' for i in range(0, len(df_reshaped.columns))]
                        df_reshaped.index = [idx_temp]
                        test_features_temp = pd.concat([test_features_temp, df_reshaped], axis=0)
                    test_features = pd.concat([test_features, test_features_temp], axis=1)

            train_features_temp = train_features.copy()
            train_features = train_features_temp.sample(frac=0.9, random_state=seed_start)
            test_features = train_features_temp.copy().drop(train_features.index)
            idx = pd.concat([train_features, test_features]).index
            train_dataset = train_dataset[time_frame_size:].reset_index(drop=True)
            train_labels = train_dataset.loc[train_features.index]['labels']
            test_labels = train_dataset.loc[test_features.index]['labels']
        else:
            if models_all.loc[i, 'include_flights']:
                train_features = train_dataset.copy().drop(['labels'], axis=1)
                test_features = test_dataset.copy().drop(['labels'], axis=1)
            else:
                train_features = train_dataset.copy().drop(['labels', 'Flights'], axis=1)
                test_features = test_dataset.copy().drop(['labels', 'Flights'], axis=1)

            train_labels = train_dataset['labels'].dropna()
            test_labels = test_dataset['labels'].dropna()

        if input_size != train_features.shape[1]:
            print(f'Model {i} has input size {input_size} but train features shape is {train_features.shape[1]}.')
            input_size = train_features.shape[1]
        models_all.loc[i, 'input_size'] = input_size
        
        # Check if model already in Summary
        workbook = openpyxl.load_workbook(summary_path)
        sheet = workbook.active
        model_in_summary = False
        less_seeds_row = 0
        less_seeds_best = seed_start
        less_seeds = False
        n = 0
        for row in sheet.iter_rows(min_row=1, values_only=True):
            model_in_summary = row[1] == models_all.loc[i, 'name'] and row[2] == models_all.loc[i, 'lookahead'] and row[3] == models_all.loc[i, 'data_start']
            model_in_summary = model_in_summary and row[4] == models_all.loc[i, 'data_end'] and row[5] == models_all.loc[i, 'learning_rate'] 
            model_in_summary = model_in_summary and row[8] == models_all.loc[i, 'file_name'] and row[9] == models_all.loc[i, 'time_frame_size']
            model_in_summary = model_in_summary and row[10] == models_all.loc[i, 'patience'] and row[11] == models_all.loc[i, 'min_delta']
            model_in_summary = model_in_summary and row[12] == input_size
            model_in_summary = model_in_summary and row[18] == models_all.loc[i, 'season_features_size']
            model_in_summary = model_in_summary and row[19] == models_all.loc[i, 'feature_threshold']
            model_in_summary = model_in_summary and row[20] == models_all.loc[i, 'learning_rate_scheduler']
            model_in_summary = model_in_summary and row[21] == models_all.loc[i, 'include_flights']
            model_in_summary = model_in_summary and row[22] == models_all.loc[i, 'num_roll_features']
            model_in_summary = model_in_summary and row[23] == models_all.loc[i, 'time_frame_features_size']
            model_in_summary = model_in_summary and row[24] == models_all.loc[i, 'time_frame_flights_size']
            model_in_summary = model_in_summary and row[25] == models_all.loc[i, 'average_window_size']
            model_in_summary = model_in_summary and row[27] == models_all.loc[i, 'average_season_features']
            model_in_summary = model_in_summary and row[28] == -1 and row[29] == -1
            model_in_summary = model_in_summary and row[33] == models_all.loc[i, 'gdp_size']
            model_in_summary = model_in_summary and row[34] == models_all.loc[i, 'add_timestamp']

            if model_in_summary and row[14] is not None and not isinstance(row[14], str) and row[14] < num_seeds:
                less_seeds_row = n
                less_seeds_best = int(row[13])  # Best seed
                less_seeds = True
                model_in_summary = False            
            n += 1
        if model_in_summary:
            print('something went wrong')
            continue
        workbook.close()
        best_history = None
        best_seed = None
        if models_all.loc[i, 'average_window_size'] > 0:
            train_features = movingAverage(train_features, models_all.loc[i, 'average_window_size'], models_all.loc[i, 'average_season_features'])
            train_labels_train = movingAverage(train_labels, models_all.loc[i, 'average_window_size'], models_all.loc[i, 'average_season_features'])
        else:
            train_labels_train = train_labels
        
        print(f'Shape train labels: {train_labels.shape}\nShape train features: {train_features.shape}')
        print(f'Time: {datetime.now().strftime('%H:%M')}')
        with tqdm(total=num_seeds-(less_seeds_best-seed_start), desc=f'Training model {i} of {len(models_all)-1}') as pbar_inner:
            for j in range(less_seeds_best-seed_start,num_seeds):
                with contextlib.redirect_stdout(dummy_file), contextlib.redirect_stderr(dummy_file):
                    set_random_seed(seed_start + j)
                    models_all.loc[i, 'model'] = models_all.loc[i, 'model_func'](train_features.shape[1])
                    models_all.loc[i, 'model'].compile(optimizer=tf.optimizers.Adam(learning_rate=models_all.loc[i, 'learning_rate']), loss='mean_absolute_percentage_error')
                    lr_schedule = tf.keras.callbacks.LearningRateScheduler(scheduler)
                    early_stopping = tf.keras.callbacks.EarlyStopping(
                        monitor='val_loss',
                        patience=models_all.loc[i, 'patience'], # Number of epochs to wait for improvement
                        min_delta=models_all.loc[i, 'min_delta'],  # Minimum change to qualify as an improvement
                        restore_best_weights=True
                    )
                    if models_all.loc[i, 'learning_rate_scheduler']:
                        history = models_all.loc[i, 'model'].fit(
                            train_features, train_labels_train,
                            epochs=max_epochs,
                            verbose=0,
                            validation_split=0.2,
                            callbacks=[early_stopping, lr_schedule]
                        )
                    else:
                        history = models_all.loc[i, 'model'].fit(
                            train_features, train_labels_train,
                            epochs=max_epochs,
                            verbose=0,
                            validation_split=0.2,
                            callbacks=[early_stopping]
                        )
                    if early_stopping.stopped_epoch == max_epochs - 1:
                        best_epoch = max_epochs - 1
                    else:
                        best_epoch = early_stopping.stopped_epoch - early_stopping.patience
                    if best_history is None or history.history['val_loss'][best_epoch] < best_history.history['val_loss'][best_best_epoch]:
                        best_history = history
                        best_model = models_all.loc[i, 'model']
                        best_seed = seed_start + j
                        best_best_epoch = best_epoch
                pbar_inner.update(1)
        models_all.loc[i, 'train_loss'] = best_history.history['loss'][best_best_epoch]
        models_all.loc[i, 'val_loss'] = best_history.history['val_loss'][best_best_epoch]
        models_all.loc[i, 'best_seed'] = best_seed
        models_all.loc[i, 'model'] = best_model
        print(f'Model {i} best epoch = {best_best_epoch}.')
        print(f'Model stopped at {early_stopping.stopped_epoch}.')
        del best_model
        gc.collect()

        actual_values_normalized = np.concatenate([train_labels, test_labels])
        actual_values = np.round(scalers_dict['Flights'].inverse_transform(actual_values_normalized.reshape(-1, 1)).flatten())

        test_values = np.round(scalers_dict['Flights'].inverse_transform(test_labels.values.reshape(-1, 1)).flatten())

        # Predict
        with contextlib.redirect_stdout(dummy_file), contextlib.redirect_stderr(dummy_file):
            predicted_values = np.round(scalers_dict['labels'].inverse_transform(models_all.loc[i, 'model'].predict(train_features).reshape(-1, 1)).flatten())

        # Compute MAPE
        min_len = min(len(actual_values), len(predicted_values)) - 1
        mape = np.mean(np.abs((actual_values[:min_len] - predicted_values[:min_len]) / actual_values[:min_len])) * 100

        # Compute HMAPE
        hmape = 1/(min_len) * np.sum(np.abs(actual_values[:min_len] - predicted_values[:min_len]) / 
                                   np.sqrt(np.abs(actual_values[:min_len] - predicted_values[:min_len]) + 1))
        
        # Test data
        with contextlib.redirect_stdout(dummy_file), contextlib.redirect_stderr(dummy_file):
            predicted_values_test = np.round(scalers_dict['labels'].inverse_transform(models_all.loc[i, 'model'].predict(test_features).reshape(-1, 1)).flatten())

        mape_test = np.mean(np.abs((test_labels - predicted_values_test) / test_labels)) * 100
        hmape_test = 1/(len(predicted_values_test)) * np.sum(np.abs(test_values[:min_len] - predicted_values_test) /
                                                        np.sqrt(np.abs(test_values - predicted_values_test) + 1))


        # Save model in summary
        workbook = openpyxl.load_workbook(summary_path)
        sheet = workbook.active
        # Write data to the next row    

        # Add buffered rows to the sheet
        for row_data in rows_buffer:
            # Save model in summary
            if row_data['less_seeds']:
                last_row = row_data['less_seeds_row']
            else:
                # Find the last row with data in column 1
                last_row = 0
                for row in range(sheet.max_row, 0, -1):
                    if sheet.cell(row=row, column=1).value is not None:
                        last_row = row
                        break
            sheet.cell(row=last_row + 1, column=1, value=row_data['id'])
            sheet.cell(row=last_row + 1, column=2, value=row_data['name'])
            sheet.cell(row=last_row + 1, column=3, value=row_data['lookahead'])
            sheet.cell(row=last_row + 1, column=4, value=row_data['data_start'])
            sheet.cell(row=last_row + 1, column=5, value=row_data['data_end'])
            sheet.cell(row=last_row + 1, column=6, value=row_data['learning_rate'])
            sheet.cell(row=last_row + 1, column=7, value=row_data['train_loss'])
            sheet.cell(row=last_row + 1, column=8, value=row_data['val_loss'])
            sheet.cell(row=last_row + 1, column=9, value=row_data['file_name'])
            sheet.cell(row=last_row + 1, column=10, value=row_data['time_frame_size'])
            sheet.cell(row=last_row + 1, column=11, value=row_data['patience'])
            sheet.cell(row=last_row + 1, column=12, value=row_data['min_delta'])
            sheet.cell(row=last_row + 1, column=13, value=row_data['input_size'])
            sheet.cell(row=last_row + 1, column=14, value=row_data['best_seed'])
            sheet.cell(row=last_row + 1, column=15, value=row_data['num_seeds'])
            sheet.cell(row=last_row + 1, column=16, value=row_data['time'])
            sheet.cell(row=last_row + 1, column=16).number_format = 'DD.MM.YYYY HH:MM:SS'
            sheet.cell(row=last_row + 1, column=19, value=row_data['season_features_size'])
            sheet.cell(row=last_row + 1, column=20, value=row_data['feature_threshold'])
            sheet.cell(row=last_row + 1, column=21, value=row_data['learning_rate_scheduler'])
            sheet.cell(row=last_row + 1, column=22, value=row_data['include_flights'])
            sheet.cell(row=last_row + 1, column=23, value=row_data['num_roll_features'])
            sheet.cell(row=last_row + 1, column=24, value=row_data['time_frame_features_size'])
            sheet.cell(row=last_row + 1, column=25, value=row_data['time_frame_flights_size'])
            sheet.cell(row=last_row + 1, column=26, value=row_data['average_window_size'])
            sheet.cell(row=last_row + 1, column=28, value=row_data['average_season_features'])
            sheet.cell(row=last_row + 1, column=29, value=-1)
            sheet.cell(row=last_row + 1, column=30, value=-1)
            sheet.cell(row=last_row + 1, column=34, value=row_data['gdp_size'])
            sheet.cell(row=last_row + 1, column=35, value=row_data['add_timestamp'])
            sheet.cell(row=last_row + 1, column=27, value=row_data['mape'])
            sheet.cell(row=last_row + 1, column=31, value=row_data['hmape'])
            sheet.cell(row=last_row + 1, column=32, value=row_data['mape_test'])
            sheet.cell(row=last_row + 1, column=33, value=row_data['hmape_test'])
        if less_seeds:
            last_row = less_seeds_row
        else:
            # Find the last row with data in column 1
            last_row = 0
            for row in range(sheet.max_row, 0, -1):
                if sheet.cell(row=row, column=1).value is not None:
                    last_row = row
                    break
        # Find the maximum ID in column 1
        max_id = 0
        for row in sheet.iter_rows(min_row=1, values_only=True):
            if row[0] is not None and isinstance(row[0], int):
                max_id = max(max_id, row[0])
        id = max_id + 1
        ids.append(id)
        sheet.cell(row=last_row + 1, column=1, value=id)
        sheet.cell(row=last_row + 1, column=2, value=models_all.loc[i, 'name'])
        sheet.cell(row=last_row + 1, column=3, value=models_all.loc[i, 'lookahead'])
        sheet.cell(row=last_row + 1, column=4, value=models_all.loc[i, 'data_start'])
        sheet.cell(row=last_row + 1, column=5, value=models_all.loc[i, 'data_end'])
        sheet.cell(row=last_row + 1, column=6, value=models_all.loc[i, 'learning_rate'])
        sheet.cell(row=last_row + 1, column=7, value=models_all.loc[i, 'train_loss'])
        sheet.cell(row=last_row + 1, column=8, value=models_all.loc[i, 'val_loss'])
        sheet.cell(row=last_row + 1, column=9, value=models_all.loc[i, 'file_name'])
        sheet.cell(row=last_row + 1, column=10, value=models_all.loc[i, 'time_frame_size'])
        sheet.cell(row=last_row + 1, column=11, value=models_all.loc[i, 'patience'])
        sheet.cell(row=last_row + 1, column=12, value=models_all.loc[i, 'min_delta'])
        sheet.cell(row=last_row + 1, column=13, value=input_size)
        sheet.cell(row=last_row + 1, column=14, value=best_seed)
        sheet.cell(row=last_row + 1, column=15, value=num_seeds)
        current_time = datetime.now()
        sheet.cell(row=last_row + 1, column=16, value=current_time)
        sheet.cell(row=last_row + 1, column=16).number_format = 'DD.MM.YYYY HH:MM:SS'
        sheet.cell(row=last_row + 1, column=19, value=models_all.loc[i, 'season_features_size'])
        sheet.cell(row=last_row + 1, column=20, value=models_all.loc[i, 'feature_threshold'])
        sheet.cell(row=last_row + 1, column=21, value=models_all.loc[i, 'learning_rate_scheduler'])
        sheet.cell(row=last_row + 1, column=22, value=models_all.loc[i, 'include_flights'])
        sheet.cell(row=last_row + 1, column=23, value=models_all.loc[i, 'num_roll_features'])
        sheet.cell(row=last_row + 1, column=24, value=models_all.loc[i, 'time_frame_features_size'])
        sheet.cell(row=last_row + 1, column=25, value=models_all.loc[i, 'time_frame_flights_size'])
        sheet.cell(row=last_row + 1, column=26, value=models_all.loc[i, 'average_window_size'])
        sheet.cell(row=last_row + 1, column=28, value=models_all.loc[i, 'average_season_features'])
        sheet.cell(row=last_row + 1, column=29, value=-1)
        sheet.cell(row=last_row + 1, column=30, value=-1)
        sheet.cell(row=last_row + 1, column=34, value=models_all.loc[i, 'gdp_size'])
        sheet.cell(row=last_row + 1, column=35, value=models_all.loc[i, 'add_timestamp'])
        sheet.cell(row=last_row + 1, column=27, value=mape)
        sheet.cell(row=last_row + 1, column=31, value=hmape)
        sheet.cell(row=last_row + 1, column=32, value=mape_test)
        sheet.cell(row=last_row + 1, column=33, value=hmape_test)

        if dump_train:
            dump(train_features, models_path + f'train_features_{id}.joblib')
        # Save model
        path = models_path + f'{id}_model.joblib'
        if not os.path.exists(path):
            dump(models_all.loc[i, 'model'], path)
        else:
            print(f"Model file {path} already exists.")
            timestamp = datetime.now().strftime("%d-%m-%Y_%H_%M_%S")
            path = models_path + f'{id}_model_{timestamp}.joblib'
            dump(models_all.loc[i, 'model'], path)

        # Save history
        path = models_path + f'{id}_history.joblib'
        if not os.path.exists(path):
            dump(best_history, path)
        else:
            print(f"History file {path} already exists.")
            timestamp = datetime.now().strftime("%d-%m-%Y_%H_%M_%S")
            path = models_path + f'{id}_history_{timestamp}.joblib'
            dump(best_history, path)

        # Save features
        path = models_path + f'{id}_features_original.joblib'
        if not os.path.exists(path):
            dump(features_original, path)
        else:
            print(f"Features file {path} already exists.")
            timestamp = datetime.now().strftime("%d-%m-%Y_%H_%M_%S")
            path = models_path + f'{id}_features_original_{timestamp}.joblib'
            dump(features_original, path)

        # Save actual values
        path = models_path + f'{id}_actual_values.joblib'
        
        # Sort back to original order
        actual_values = actual_values[np.argsort(idx)]
        if not os.path.exists(path):
            dump(actual_values, path)
        else:
            print(f"Actual values file {path} already exists.")
            timestamp = datetime.now().strftime("%d-%m-%Y_%H_%M_%S")
            path = models_path + f'{id}_actual_values_{timestamp}.joblib'
            dump(actual_values, path)
        
        # Save scalers
        path = models_path + f'{id}_scalers.joblib'
        if not os.path.exists(path):
            dump(scalers_dict, path)
        else:
            print(f"Scalers file {path} already exists.")
            timestamp = datetime.now().strftime("%d-%m-%Y_%H_%M_%S")
            path = models_path + f'{id}_scalers_{timestamp}.joblib'
            dump(scalers_dict, path)
        
        # Save test features
        path = models_path + f'{id}_test_features.joblib'
        if not os.path.exists(path):
            dump(test_features, path)
        else:
            print(f"Test features file {path} already exists.")
            timestamp = datetime.now().strftime("%d-%m-%Y_%H_%M_%S")
            path = models_path + f'{id}_test_features_{timestamp}.joblib'
            dump(test_features, path)

        # Save test labels
        path = models_path + f'{id}_test_labels.joblib'
        if not os.path.exists(path):
            dump(test_labels, path)
        else:
            print(f"Test labels file {path} already exists.")
            timestamp = datetime.now().strftime("%d-%m-%Y_%H_%M_%S")
            path = models_path + f'{id}_test_labels_{timestamp}.joblib'
            dump(test_labels, path)

        # Free memory
        models_all.loc[i, 'model'] = None
        train_dataset_list[i] = None
        test_dataset_list[i] = None
        idx_list[i] = None
        features_original_list[i] = None
        history = None
        best_history = None
        gc.collect()

        # Try to save the Excel file
        try:
            workbook.save(summary_path)
            rows_buffer = list()
            print("File saved successfully.")
            if restart_script and i > 30:
                print("Restarting the script.")
                subprocess.Popen(["python", "start_script_train.py"])
                return
        except PermissionError:
            workbook = openpyxl.load_workbook(summary_path)
            sheet = workbook.active
            rows_buffer.append({
                'id': id,
                'name': models_all.loc[i, 'name'],
                'lookahead': models_all.loc[i, 'lookahead'],
                'data_start': models_all.loc[i, 'data_start'],
                'data_end': models_all.loc[i, 'data_end'],
                'learning_rate': models_all.loc[i, 'learning_rate'],
                'train_loss': models_all.loc[i, 'train_loss'],
                'val_loss': models_all.loc[i, 'val_loss'],
                'file_name': models_all.loc[i, 'file_name'],
                'time_frame_size': models_all.loc[i, 'time_frame_size'],
                'patience': models_all.loc[i, 'patience'],
                'min_delta': models_all.loc[i, 'min_delta'],
                'input_size': input_size,
                'best_seed': best_seed,
                'num_seeds': num_seeds,
                'less_seeds': less_seeds,
                'less_seeds_row': less_seeds_row,
                'time': current_time,
                'season_features_size': models_all.loc[i, 'season_features_size'],
                'feature_threshold': models_all.loc[i, 'feature_threshold'],
                'learning_rate_scheduler': models_all.loc[i, 'learning_rate_scheduler'],
                'include_flights': models_all.loc[i, 'include_flights'],
                'num_roll_features': models_all.loc[i, 'num_roll_features'],
                'time_frame_features_size': models_all.loc[i, 'time_frame_features_size'],
                'time_frame_flights_size': models_all.loc[i, 'time_frame_flights_size'],
                'average_window_size': models_all.loc[i, 'average_window_size'],
                'average_season_features': models_all.loc[i, 'average_season_features'],
                'gdp_size': models_all.loc[i, 'gdp_size'],
                'add_timestamp': models_all.loc[i, 'add_timestamp'],
                'mape': mape,
                'hmape': hmape,
                'mape_test': mape_test,
                'hmape_test': hmape_test
            })
            print(f"The file is open. Row with ID {id} saved in buffer.")
            print(f'Buffer size: {len(rows_buffer)}')
        elapsed_time = time.time() - start_time
        # Estimate remaining time
        iterations_left = total_iterations - (i + 1)
        time_per_iteration = elapsed_time / (i + 1)
        estimated_time_left = iterations_left * time_per_iteration
            
        # Convert elapsed time to hh:mm format
        elapsed_hours, elapsed_minutes = divmod(elapsed_time // 60, 60)
        # Convert estimated time left to hh:mm format
        estimated_hours, estimated_minutes = divmod(estimated_time_left // 60, 60)
        # Print elapsed time and estimated time left in hh:mm format
        print(f"Elapsed time: {int(elapsed_hours):02d}:{int(elapsed_minutes):02d}")
        print(f"Estimated time left: {int(estimated_hours):02d}:{int(estimated_minutes):02d}")
        # Calculate estimated time of arrival (ETA)
        eta = datetime.now() + timedelta(seconds=estimated_time_left)
        print(f"ETA: {eta.strftime('%d.%m.%Y %H:%M:%S')}\n")
            


    if len(rows_buffer) > 0:
        print('All models finished.')
        print('Waiting for the file to be closed...')

    while len(rows_buffer) > 0:
        time.sleep(3)
        # Add buffered rows to the sheet
        for row_data in rows_buffer:
            # Save model in summary
            if row_data['less_seeds']:
                last_row = row_data['less_seeds_row']
            else:
                # Find the last row with data in column 1
                last_row = 0
                for row in range(sheet.max_row, 0, -1):
                    if sheet.cell(row=row, column=1).value is not None:
                        last_row = row
                        break
            sheet.cell(row=last_row + 1, column=1, value=row_data['id'])
            sheet.cell(row=last_row + 1, column=2, value=row_data['name'])
            sheet.cell(row=last_row + 1, column=3, value=row_data['lookahead'])
            sheet.cell(row=last_row + 1, column=4, value=row_data['data_start'])
            sheet.cell(row=last_row + 1, column=5, value=row_data['data_end'])
            sheet.cell(row=last_row + 1, column=6, value=row_data['learning_rate'])
            sheet.cell(row=last_row + 1, column=7, value=row_data['train_loss'])
            sheet.cell(row=last_row + 1, column=8, value=row_data['val_loss'])
            sheet.cell(row=last_row + 1, column=9, value=row_data['file_name'])
            sheet.cell(row=last_row + 1, column=10, value=row_data['time_frame_size'])
            sheet.cell(row=last_row + 1, column=11, value=row_data['patience'])
            sheet.cell(row=last_row + 1, column=12, value=row_data['min_delta'])
            sheet.cell(row=last_row + 1, column=13, value=row_data['input_size'])
            sheet.cell(row=last_row + 1, column=14, value=row_data['best_seed'])
            sheet.cell(row=last_row + 1, column=15, value=row_data['num_seeds'])
            sheet.cell(row=last_row + 1, column=16, value=row_data['time'])
            sheet.cell(row=last_row + 1, column=16).number_format = 'DD.MM.YYYY HH:MM:SS'
            sheet.cell(row=last_row + 1, column=19, value=row_data['season_features_size'])
            sheet.cell(row=last_row + 1, column=20, value=row_data['feature_threshold'])
            sheet.cell(row=last_row + 1, column=21, value=row_data['learning_rate_scheduler'])
            sheet.cell(row=last_row + 1, column=22, value=row_data['include_flights'])
            sheet.cell(row=last_row + 1, column=23, value=row_data['num_roll_features'])
            sheet.cell(row=last_row + 1, column=24, value=row_data['time_frame_features_size'])
            sheet.cell(row=last_row + 1, column=25, value=row_data['time_frame_flights_size'])
            sheet.cell(row=last_row + 1, column=26, value=row_data['average_window_size'])
            sheet.cell(row=last_row + 1, column=28, value=row_data['average_season_features'])
            sheet.cell(row=last_row + 1, column=29, value=-1)
            sheet.cell(row=last_row + 1, column=30, value=-1)
            sheet.cell(row=last_row + 1, column=34, value=row_data['gdp_size'])
            sheet.cell(row=last_row + 1, column=35, value=row_data['add_timestamp'])
            sheet.cell(row=last_row + 1, column=27, value=row_data['mape'])
            sheet.cell(row=last_row + 1, column=31, value=row_data['hmape'])
            sheet.cell(row=last_row + 1, column=32, value=row_data['mape_test'])
            sheet.cell(row=last_row + 1, column=33, value=row_data['hmape_test'])
        try:
            workbook.save(summary_path)
            rows_buffer = list()
            print("File saved successfully.")
            break
        except PermissionError:
            workbook = openpyxl.load_workbook(summary_path)
            sheet = workbook.active
            print(f"The file {summary_path} is open. Waiting for it to be closed...")

    print(f'All models saved in Summary at {datetime.now().strftime("%d.%m.%Y %H:%M:%S")}')
    return ids





if __name__ == '__main__':
    main()