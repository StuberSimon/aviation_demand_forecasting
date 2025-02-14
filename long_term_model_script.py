class DummyFile(object):
    def write(self, x): pass
    def flush(self): pass

import os
# Suppress TensorFlow warnings
os.environ['TF_CPP_MIN_LOG_LEVEL'] = '3'
import numpy as np
import pandas as pd
from joblib import dump, load
import tensorflow as tf
from tensorflow import keras
from keras import layers
from tqdm import tqdm
import random
from datetime import datetime, timedelta
from tsfresh.utilities.dataframe_functions import roll_time_series, impute
from sklearn.preprocessing import MinMaxScaler
import openpyxl
import time
import contextlib
import shutil
import gc
import matplotlib.pylab as plt
from tsfresh import extract_features
dummy_file = DummyFile()


def movingAverage(array, size, average_season_features, first=True):
    # Convert array to DataFrame if it is not already one
    if not isinstance(array, pd.DataFrame):
        idx_temp = array.index
        array = pd.DataFrame(array)
        array.index = idx_temp
    if first:
        result = np.zeros(array.shape, dtype=float)
    else:
        result = np.zeros((1, array.shape[1]), dtype=float)
    for col in range(array.shape[1]):
        if first:
            if (not average_season_features) and (array.columns[col][:8] in ['3_months', '4_months', '6_months'] or array.columns[col][:9] in ['12_months', '17_months', '54_months']):
                    result[:, col] = array.iloc[:, col]
            else:
                for i in range(array.shape[0]):
                    if i < size + 1:
                        result[i, col] = np.mean(array.iloc[:i+1, col])
                    elif i >= array.shape[0] - size:
                        result[i, col] = np.mean(array.iloc[i-size:, col])
                    else:
                        result[i, col] = np.mean(array.iloc[i-size:i+1, col])
        else: # If not first
            if not average_season_features and (array.columns[col][:8] in ['3_months', '4_months', '6_months'] or array.columns[col][:9] in ['12_months', '17_months', '54_months']):
                result[0, col] = array.iloc[-1, col]
            else:
                result[0, col] = np.mean(array.iloc[-size-1:, col]) # Just last row
    if first:
        result_df = pd.DataFrame(result, columns=array.columns, index=array.index)
    else:
        result_df = pd.DataFrame(result, columns=array.columns)
    return result_df

def main():
    # Base path for the extracted features with '/'
    base_path = '01_extracted_features/'
    models_path = '02_models/'
    summary_path = 'Summary.xlsx'
    plot_path = '05_long_term_plots/'
    preds_path = '06_long_term_predictions/'

    model_ids = []
    end_dates_list = ['2050-12']

    season_features = load(base_path + 'season_features_long_term_3_4_6_12_17_54.joblib')
    feature_importance = load(base_path + '1290_feature_importance_df.joblib')
    gdp_features = load(base_path + 'gdp_features_v1.joblib').reset_index(drop=True)

    long_term_forecast(model_ids, end_dates_list, base_path, models_path, summary_path, plot_path, preds_path, season_features, feature_importance, gdp_features)

    return
    
def long_term_forecast(model_ids=None, end_dates_list=None, base_path=None, models_path=None, summary_path=None, plot_path=None, preds_path=None, season_features=None,
                       feature_importance=None, gdp_features=None):
    
    """
    Forecast long term values for the given models and end dates.

    Args:
        model_ids (list): List of model IDs to forecast.
        end_dates_list (list): List of end dates to forecast.
        base_path (str): Path to the extracted features.
        models_path (str): Path to the models.
        summary_path (str): Path to the summary file.
        plot_path (str): Path to save the plots.
        preds_path (str): Path to save the predicted values.
        season_features (pd.DataFrame): Seasonal features.
        feature_importance (pd.DataFrame): Feature importance.
        gdp_features (pd.DataFrame): GDP features.

    Returns:
        list: List of forecasted model IDs.
    """

    print('Starting long term forecasting...')
    
    # Set missing parameters to default
    model_ids = model_ids or []
    end_dates_list = end_dates_list or ['2050-12']
    base_path = base_path or '01_extracted_features/'
    models_path = models_path or '02_models/'
    summary_path = summary_path or 'Summary.xlsx'
    plot_path = plot_path or '05_long_term_plots/'
    preds_path = preds_path or '06_long_term_predictions/'
    season_features = season_features or load(base_path + 'season_features_long_term_3_4_6_12_17_54.joblib')
    feature_importance = feature_importance or load(base_path + '1290_feature_importance_df.joblib')
    gdp_features = gdp_features or load(base_path + 'gdp_features_v1.joblib').reset_index(drop=True)

    # Backup Summary
    # Get the base name of the source file
    base_name = os.path.basename(summary_path)
    # Get the file name and extension
    file_name, file_ext = os.path.splitext(base_name)
    # Get the current date
    current_date = datetime.now().strftime('%d.%m.%Y_%H-%M-%S')	
    # Create the new file name with date
    new_file_name = f"{file_name}_{current_date}{file_ext}"
    # Create the destination path
    dest_path = os.path.join('04_Summary_backups', new_file_name)
    # Copy the file to the new destination
    shutil.copy(summary_path, dest_path)

    # Build df for all models
    models_all = pd.DataFrame(
        columns=['id', 'ref_id', 'name', 'lookahead', 'data_start', 'data_end', 'learning_rate',
                 'file_name', 'model', 'time_frame_size', 'patience', 'min_delta', 'input_size',
                'season_features_size', 'feature_threshold', 'learning_rate_scheduler', 'include_flights',
                'num_roll_features', 'time_frame_features_size', 'time_frame_flights_size', 'average_window_size',
                'average_season_features', 'num_epochs', 'end_date', 'gdp_size'],
        index=range(len(model_ids) * len(end_dates_list))
    )

    for i in range(len(models_all)):
        models_all.loc[i, 'id'] = model_ids[i % len(model_ids)]
        models_all.loc[i, 'end_date'] = end_dates_list[i // len(model_ids) % len(end_dates_list)]

    workbook = openpyxl.load_workbook(summary_path)
    sheet = workbook.active
    
    for i in range(len(models_all)):
        row = None
        for r in sheet.iter_rows(min_row=0, max_row=sheet.max_row, values_only=True):
            if r[0] == models_all.loc[i, 'id']:
                row = r
                break
        if row is None:
            print(f'Model {models_all.loc[i, 'id']} not found in Summary.')
            continue
        models_all.loc[i, 'ref_id'] = row[28]
        models_all.loc[i, 'lookahead'] = row[2]
        models_all.loc[i, 'data_start'] = row[3]
        models_all.loc[i, 'data_end'] = row[4]
        models_all.loc[i, 'file_name'] = row[8]
        models_all.loc[i, 'name'] = row[1]
        models_all.loc[i, 'learning_rate'] = row[5]
        models_all.loc[i, 'time_frame_size'] = row[9]
        models_all.loc[i, 'patience'] = row[10]
        models_all.loc[i, 'min_delta'] = row[11]
        models_all.loc[i, 'input_size'] = row[12]
        models_all.loc[i, 'season_features_size'] = row[18]
        models_all.loc[i, 'feature_threshold'] = row[19]
        models_all.loc[i, 'learning_rate_scheduler'] = row[20]
        models_all.loc[i, 'include_flights'] = row[21]
        models_all.loc[i, 'num_roll_features'] = row[22]
        models_all.loc[i, 'time_frame_features_size'] = row[23]
        models_all.loc[i, 'time_frame_flights_size'] = row[24]
        models_all.loc[i, 'average_window_size'] = row[25]
        models_all.loc[i, 'average_season_features'] = row[27]
        models_all.loc[i, 'num_epochs'] = row[29]
        models_all.loc[i, 'window_size'] = int(row[8][-6:-4])
        models_all.loc[i, 'gdp_size'] = row[33]
        models_all.loc[i, 'add_timestamp'] = row[34]

    models_all = models_all[models_all['name'].notna()].reset_index(drop=True)
    # Save models_all in csv
    models_all.to_excel('sbs_models_all.xlsx', index=True)
    print('Models_all saved.')

    start_time = time.time()
    total_iterations = len(models_all)
    n_skipped = 0
    ids = list()

    for i in range(len(models_all)):
        try:
            if os.path.exists(plot_path + f'{models_all.loc[i, 'id']}_long_term_plot_{models_all.loc[i, 'end_date']}.png') and os.path.exists(preds_path + f'{models_all.loc[i, 'id']}_predicted_values_{models_all.loc[i, 'end_date']}.joblib'):
                print(f"Plot and pred for model {i}/{len(models_all)-1} with id {models_all.loc[i, 'id']} already exists.")
                n_skipped += 1
                continue
            n_steps = 0
            num_preds_offset = 0
            if models_all.loc[i, 'ref_id'] >= 0:
                models_all.loc[i, 'data_end'] = 321 # Start SBS at end of data
                num_preds_offset = 1
            data_end_original = models_all.loc[i, 'data_end']
            #models_all.loc[i, 'data_end'] += 1
            predicted_values = np.array([])
            train_features_saved_avg = pd.DataFrame()
            new_features = pd.DataFrame()
            features_calculated = pd.DataFrame()
            date_range = pd.date_range(start='1997-09', end=models_all.loc[i, 'end_date'], freq='ME').strftime('%Y-%m')
            num_preds = len(date_range) - (models_all.loc[i, 'data_end'] - models_all.loc[i, 'data_start'])
            features_raw_original = pd.read_csv(base_path + models_all.loc[i, 'file_name']).drop(['YearMonth'], axis=1)
            pbar_outer = None

            while len(predicted_values) < num_preds:
                if n_steps == 1:
                    pbar_outer = tqdm(total=num_preds-num_preds_offset, desc=f'Forecasting for model {i}/{len(models_all)-1}')

                if n_steps > 0:
                    features_raw = features_raw_original.copy().iloc[models_all.loc[i, 'data_start']:data_end_original] # Actual values are always the same
                    actual_values = features_raw_original['Flights'].copy()
                    window = pd.DataFrame(columns=['id', 'Time', 'Flights'])
                    window['Flights'] = pd.DataFrame(np.concatenate([features_raw['Flights'], predicted_values], axis=0))[int(-models_all.loc[i, 'window_size']-1):]
                    window['id'] = 'A'
                    window['Time'] = window.index
                    window = window.reset_index(drop=True)
                    with contextlib.redirect_stdout(dummy_file), contextlib.redirect_stderr(dummy_file):
                        new_features = extract_features(window, column_id="id", column_sort="Time", 
                                                        column_value="Flights", impute_function=impute, show_warnings=False)
                    new_features['Flights'] = window['Flights'].iloc[-1]
                    features_calculated = pd.concat([features_calculated, new_features], axis=0).reset_index(drop=True)
                    features_raw = pd.concat([features_raw, features_calculated], axis=0).reset_index(drop=True)
                else:
                    features_raw = features_raw_original.copy().iloc[models_all.loc[i, 'data_start']:models_all.loc[i, 'data_end']]
                    actual_values = features_raw_original['Flights'].copy()

                important_features = feature_importance[feature_importance['cumulative_importance'] <= models_all.loc[i, 'feature_threshold']]
                important_features = important_features.sort_values(by='cumulative_importance', ascending=True)
                valid_features = [feature for feature in important_features['feature'] if feature in features_raw.columns]
                valid_features.append('Flights')
                features_raw = features_raw[valid_features]

                if models_all.loc[i, 'season_features_size'] > 0:
                    features_raw = pd.concat([features_raw, season_features.iloc[models_all.loc[i, 'data_start']:models_all.loc[i, 'data_end']]], axis=1) 
                if models_all.loc[i, 'gdp_size'] > 0:
                    features_raw = pd.concat([gdp_features.iloc[models_all.loc[i, 'data_start']:models_all.loc[i, 'data_end'],:models_all.loc[i, 'gdp_size']], features_raw], axis=1)
                if models_all.loc[i, 'add_timestamp'] == 1:
                    features_raw['Timestamp'] = features_raw.index

                train_dataset = features_raw.copy()

                # Normalize the data
                if n_steps == 0:
                    scalers_dict = load(models_path + f'{models_all.loc[i, 'id']}_scalers.joblib')
                for column in train_dataset.columns:
                    scaler = scalers_dict[column]
                    train_dataset[column] = scaler.transform(train_dataset[column].values.reshape(-1, 1))
                
                time_frame_features_size = models_all.loc[i, 'time_frame_features_size'] - 1
                time_frame_flights_size = models_all.loc[i, 'time_frame_flights_size'] - 1
                time_frame_size = models_all.loc[i, 'time_frame_size']
                num_roll_features = models_all.loc[i, 'num_roll_features']

                if time_frame_size > 0 and len(train_dataset) >= time_frame_size:
                    # Select first ten columns from train_dataset and append 'Flights' if not already in
                    if time_frame_features_size > 0:
                        selected_columns = train_dataset.copy().drop(['Flights'], axis=1).columns[:num_roll_features].tolist()
                    else:
                        selected_columns = []
                    # Train features
                    if n_steps == 0: # First rolling
                        train_features_saved = pd.DataFrame()
                        if models_all.loc[i, 'include_flights'] and time_frame_flights_size == 0:
                            train_features = train_dataset.copy().drop(columns=selected_columns)[time_frame_size:].reset_index(drop=True)
                        elif not models_all.loc[i, 'include_flights']:
                            train_features = train_dataset.copy().drop(columns=selected_columns).drop(['Flights'], axis=1)[time_frame_size:].reset_index(drop=True)
                        else:
                            selected_columns.append('Flights')
                            train_features = train_dataset.copy().drop(columns=selected_columns, axis=1)[time_frame_size:].reset_index(drop=True)
                    else: # Select only the last row
                        if models_all.loc[i, 'include_flights'] and time_frame_flights_size == 0:
                            train_features = train_dataset.copy().drop(columns=selected_columns)[time_frame_size:].reset_index(drop=True).tail(1)
                        elif not models_all.loc[i, 'include_flights']:
                            train_features = train_dataset.copy().drop(columns=selected_columns).drop(['Flights'], axis=1)[time_frame_size:].reset_index(drop=True).tail(1)
                        else:
                            selected_columns.append('Flights')
                            train_features = train_dataset.copy().drop(columns=selected_columns, axis=1)[time_frame_size:].reset_index(drop=True).tail(1)
                    
                    if n_steps == 0:
                        pbar_inner = tqdm(total=len(selected_columns), desc=f'Rolling train for model {i}/{len(models_all)-1}')
                    else:
                        pbar_inner = None
                    for feature in selected_columns:
                        if n_steps == 0:
                            train_dataset_melted = train_dataset[[feature]].copy()
                            train_dataset_melted['Timestamp'] = train_dataset.index
                            train_dataset_melted['Symbols'] = 'A'
                            # Roll time series
                            with contextlib.redirect_stdout(dummy_file), contextlib.redirect_stderr(dummy_file):
                                if feature == 'Flights' and time_frame_flights_size > 0 and models_all.loc[i, 'include_flights']:
                                    train_features_rolled = roll_time_series(train_dataset_melted, column_id='Symbols',
                                                                            column_sort='Timestamp', max_timeshift=time_frame_flights_size,
                                                                            min_timeshift=time_frame_flights_size)
                                elif feature != 'Flights' and time_frame_features_size > 0:
                                    train_features_rolled = roll_time_series(train_dataset_melted, column_id='Symbols',
                                                                            column_sort='Timestamp', max_timeshift=time_frame_features_size,
                                                                            min_timeshift=time_frame_features_size)
                                else:
                                    continue
                                # Transform rolled features
                                grouped = train_features_rolled.groupby('id')
                                dfs = {name: group for name, group in grouped}
                                train_features_temp = pd.DataFrame()
                                for key in dfs.keys():
                                    df = dfs[key].drop(['id', 'Symbols'], axis=1)
                                    df.index = df['Timestamp'].values
                                    df = df.drop(['Timestamp'], axis=1)
                                    idx_temp = df.index[0]
                                    # Transpose the DataFrame
                                    df_reshaped = df.T
                                    # Reset the index to ensure the original index is preserved
                                    df_reshaped.columns = [f'{feature}_{i}' for i in range(0, len(df_reshaped.columns))]
                                    df_reshaped.index = [idx_temp]
                                    train_features_temp = pd.concat([train_features_temp, df_reshaped], axis=0)
                                if n_steps == 0 and feature == 'Flights':
                                    if time_frame_flights_size < time_frame_features_size:
                                        train_features_temp = train_features_temp[time_frame_features_size-time_frame_flights_size:].reset_index(drop=True)
                                elif n_steps == 0 and feature != 'Flights':
                                    if time_frame_features_size < time_frame_flights_size:
                                        train_features_temp = train_features_temp[time_frame_flights_size-time_frame_features_size:].reset_index(drop=True)
                            if pbar_inner:
                                pbar_inner.update(1)
                        else: # Rolling should only output one row
                            if feature == 'Flights' and time_frame_flights_size > 0 and models_all.loc[i, 'include_flights']:
                                train_dataset_melted = train_dataset[[feature]][-time_frame_flights_size-1:].copy().reset_index(drop=True)
                            elif feature != 'Flights' and time_frame_features_size > 0:
                                train_dataset_melted = train_dataset[[feature]][-time_frame_features_size-1:].copy().reset_index(drop=True)
                            else:
                                continue
                            # Ensure the input is a NumPy array
                            arr = np.asarray(train_dataset_melted)                            
                            # Create column names
                            col_names = [f"{feature}_{i}" for i in range(arr.shape[0])]                            
                            # Convert the array to a DataFrame and transpose it
                            train_features_temp = pd.DataFrame(arr).T                            
                            # Set the column names
                            train_features_temp.columns = col_names
                            train_features_temp.index = train_features.index
                            
                        train_features = pd.concat([train_features, train_features_temp], axis=1)
                    if pbar_inner:
                        del pbar_inner
                    train_features = pd.concat([train_features_saved, train_features], axis=0)
                    train_features_saved = train_features.copy()
                else:
                    if models_all.loc[i, 'include_flights']:
                        train_features = train_dataset.copy()
                    else:
                        train_features = train_dataset.copy().drop(['Flights'], axis=1)

                models_all.loc[i, 'input_size'] = train_features.shape[1]


                if models_all['name'].loc[i][:4] == 'LSTM':
                    train_features = train_features.values.reshape(train_features.shape[0], 1, train_features.shape[1])
                if models_all.loc[i, 'average_window_size'] > 0:
                    train_features_temp = movingAverage(train_features, models_all.loc[i, 'average_window_size'], models_all.loc[i, 'average_season_features'], n_steps==0)
                    train_features = pd.concat([train_features_saved_avg, train_features_temp], axis=0).reset_index(drop=True)
                    train_features_saved_avg = train_features.copy()
                if n_steps == 0:
                    model = load(models_path + f'{models_all.loc[i, 'id']}_model.joblib')
                    train_features_original = train_features.copy()

                models_all.loc[i, 'data_end'] += 1
                # Predict
                with contextlib.redirect_stdout(dummy_file), contextlib.redirect_stderr(dummy_file):
                    predicted_value = np.round(scalers_dict['labels'].inverse_transform(model.predict(train_features.iloc[-1,:].values.reshape(1, -1))).flatten())
                predicted_values = np.append(predicted_values, predicted_value)
                n_steps += 1
                if pbar_outer:
                    pbar_outer.update(1)
            id = models_all.loc[i, 'id']

            # Original predictions
            predictions_filename = models_path + f'{models_all.loc[i, 'id']}_predicted_values.joblib'
            if models_all.loc[i, 'ref_id'] >= 0 and os.path.exists(predictions_filename): # If SBS
                predicted_values_original = load(predictions_filename)
            else:
                model_original = load(models_path + f'{models_all.loc[i, 'id']}_model.joblib')
                scalers_dict_original = load(models_path + f'{models_all.loc[i, 'id']}_scalers.joblib')
                with contextlib.redirect_stdout(dummy_file), contextlib.redirect_stderr(dummy_file):
                    predicted_values_original = np.round(scalers_dict_original['labels'].inverse_transform(model_original.predict(train_features_original).reshape(-1, 1)).flatten()) #train_features[:data_end_original-lookahead-time_frame_size-1]
            predicted_values = np.append(predicted_values_original[:-1], predicted_values)



            # Plot values
            if time_frame_size == 0:
                actual_values = actual_values[models_all.loc[i, 'data_start']+models_all.loc[i, 'lookahead']:].reset_index(drop=True)
            else:
                actual_values = actual_values[models_all.loc[i, 'data_start']+models_all.loc[i, 'lookahead']+time_frame_size+1:].reset_index(drop=True)
            plt.figure(figsize=(15, 8))
            plt.plot(actual_values, label='Actual Values')
            plt.plot(predicted_values, label='Predicted Values')
            plt.xlabel('Time')
            plt.ylabel('Number of Flights')
            title = 'Model name: ' + models_all.loc[i, 'name'] + '; Epochs: ' + str(models_all.loc[i, 'num_epochs'])
            title += '; Data points: '  + str(models_all.loc[i, 'data_start']) + '-' + str(data_end_original)
            title += f'; Learning rate: {models_all.loc[i, 'learning_rate']}'
            title += f'; Time frame size: {models_all.loc[i, 'time_frame_size']};\n'
            title += f'Data: {models_all.loc[i, 'file_name']}' + '; End date: ' + models_all.loc[i, 'end_date']
            plt.title(title)
            plt.legend()
            plt.grid(True)

            # Format x-axis
            # Generate date range starting from September 1997
            if time_frame_size == 0:
                start_date = pd.to_datetime('1997-09') + pd.DateOffset(months=models_all.loc[i, 'data_start'] + models_all.loc[i, 'lookahead'])
            else:
                start_date = pd.to_datetime('1997-09') + pd.DateOffset(months=models_all.loc[i, 'data_start'] + models_all.loc[i, 'lookahead'] + time_frame_size + 1)
            date_range = pd.date_range(start=start_date, periods=models_all.loc[i, 'data_end'] - models_all.loc[i, 'data_start'], freq='ME')
            # Extract years and set x-ticks for the start of each year
            x_ticks = [i for i, date in enumerate(date_range) if date.month == 1 and date.year % 5 == 0]
            x_labels = [date.year for date in date_range if date.month == 1 and date.year % 5 == 0]
            # Set x-axis
            plt.xticks(ticks=x_ticks, labels=x_labels)
            plt.xlim(0, len(predicted_values))  # Set x-axis limits

            plt.savefig(plot_path + f'{id}_long_term_plot_{models_all.loc[i, 'end_date']}.png', dpi=300)
            plt.close()

            # Save predicted values
            path = preds_path + f'{id}_predicted_values_{models_all.loc[i, 'end_date']}.joblib'
            if not os.path.exists(path):
                dump(predicted_values, path)
            else:
                print(f"Predicted values file {path} already exists.")
                timestamp = datetime.now().strftime("%d-%m-%Y_%H_%M_%S")
                path = preds_path + f'{id}_predicted_values_{models_all.loc[i, 'end_date']}_{timestamp}.joblib'
                dump(predicted_values, path)

            elapsed_time = time.time() - start_time
            # Estimate remaining time
            iterations_left = total_iterations - (i + 1)
            time_per_iteration = elapsed_time / (i + 1 - n_skipped)
            estimated_time_left = iterations_left * time_per_iteration
            # Convert elapsed time to hh:mm format
            elapsed_hours, elapsed_minutes = divmod(elapsed_time // 60, 60)
            # Convert estimated time left to hh:mm format
            estimated_hours, estimated_minutes = divmod(estimated_time_left // 60, 60)
            # Print elapsed time and estimated time left in hh:mm format
            print(f"Elapsed time: {int(elapsed_hours):02d}:{int(elapsed_minutes):02d}")
            print(f"Estimated time left: {int(estimated_hours):02d}:{int(estimated_minutes):02d}")
            # Calculate estimated time of arrival (ETA)
            eta = datetime.now() + timedelta(seconds=estimated_time_left)
            print(f"ETA: {eta.strftime('%d.%m.%Y %H:%M:%S')}\n")
            del pbar_outer
            ids.append(models_all.loc[i, 'id'])
        except Exception as e:
            print(f"Error for model {i}/{len(models_all)-1} with id {models_all.loc[i, 'id']}:\n{e}")
            n_skipped += 1
            continue
    print(f'Finished forecasting for {len(models_all)-n_skipped} models.')
    return ids







if __name__ == '__main__':
    main()